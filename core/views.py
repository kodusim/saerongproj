from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, Sum, Q
from collector.models import CollectedData, CrawlLog
from sources.models import DataSource
from core.models import Category, SubCategory
from core import moscom_client
from core import predictor
from core import user_store
from core import remedy_store
from core import report_store
from core import kakao_client
import logging
import json as _json

logger = logging.getLogger(__name__)


def get_crawler_status():
    """크롤러 상태 정보 가져오기"""
    from saerong.celery import app as celery_app
    from datetime import datetime

    try:
        inspect = celery_app.control.inspect()
        active_tasks = inspect.active() or {}
        reserved_tasks = inspect.reserved() or {}

        current_task = None
        is_running = False
        queue_length = 0
        queued_sources = []

        # 활성 태스크 확인
        for worker, tasks in active_tasks.items():
            for task in tasks:
                if task.get('name') == 'collector.tasks.crawl_data_source':
                    is_running = True
                    source_id = task['args'][0] if task.get('args') else None
                    time_start = task.get('time_start')

                    if source_id:
                        try:
                            source = DataSource.objects.select_related('subcategory').get(id=source_id)
                            current_task = {
                                'source_id': source_id,
                                'source_name': source.name,
                                'game_name': source.subcategory.name if source.subcategory else '',
                                'started_at': datetime.fromtimestamp(time_start) if time_start else None
                            }
                        except DataSource.DoesNotExist:
                            current_task = {
                                'source_id': source_id,
                                'source_name': '알 수 없음',
                                'game_name': '',
                                'started_at': datetime.fromtimestamp(time_start) if time_start else None
                            }
                    break

        # 대기 중인 태스크 확인
        for worker, tasks in reserved_tasks.items():
            for task in tasks:
                if task.get('name') == 'collector.tasks.crawl_data_source':
                    queue_length += 1
                    source_id = task['args'][0] if task.get('args') else None
                    if source_id:
                        try:
                            source = DataSource.objects.select_related('subcategory').get(id=source_id)
                            queued_sources.append({
                                'source_name': source.name,
                                'game_name': source.subcategory.name if source.subcategory else ''
                            })
                        except DataSource.DoesNotExist:
                            queued_sources.append({
                                'source_name': f'소스 #{source_id}',
                                'game_name': ''
                            })

        # 최근 크롤링 결과
        recent_logs = CrawlLog.objects.select_related('source', 'source__subcategory').order_by('-completed_at')[:5]
        last_crawl_results = []
        for log in recent_logs:
            last_crawl_results.append({
                'source_name': log.source.name if log.source else '알 수 없음',
                'game_name': log.source.subcategory.name if log.source and log.source.subcategory else '',
                'status': log.status,
                'items_collected': log.items_collected,
                'completed_at': log.completed_at,
                'duration_seconds': log.duration_seconds
            })

        return {
            'is_running': is_running,
            'current_task': current_task,
            'queue_length': queue_length,
            'queued_sources': queued_sources[:5],
            'total_sources': DataSource.objects.filter(is_active=True).count(),
            'last_crawl_results': last_crawl_results
        }

    except Exception as e:
        return {
            'is_running': False,
            'current_task': None,
            'queue_length': 0,
            'queued_sources': [],
            'total_sources': DataSource.objects.filter(is_active=True).count(),
            'last_crawl_results': [],
            'error': str(e)
        }


def dashboard(request):
    """메인 페이지 - 대분류 카테고리 표시"""
    categories = Category.objects.filter(is_active=True).order_by('order', 'name')

    # 크롤러 상태 정보 가져오기
    crawler_status = get_crawler_status()

    context = {
        'categories': categories,
        'crawler': crawler_status,
    }

    return render(request, 'core/home.html', context)


def category_detail(request, slug):
    """대분류 상세 - 중분류 목록 표시"""
    category = get_object_or_404(Category, slug=slug, is_active=True)

    # 검색어
    search_query = request.GET.get('search', '')

    # 중분류 목록 (활성 데이터 소스 개수 포함)
    subcategories = SubCategory.objects.filter(
        category=category,
        is_active=True
    ).annotate(
        active_sources_count=Count('data_sources', filter=Q(data_sources__is_active=True))
    ).order_by('order', 'name')

    # 검색 필터링
    if search_query:
        subcategories = subcategories.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    context = {
        'category': category,
        'subcategories': subcategories,
        'search_query': search_query,
    }

    return render(request, 'core/category_detail.html', context)


def subcategory_detail(request, slug):
    """중분류 상세 - 소분류별 데이터 표시 (탭 형식)"""
    subcategory = get_object_or_404(SubCategory, slug=slug, is_active=True)

    # 활성화된 데이터 소스들 가져오기
    data_sources = subcategory.data_sources.filter(is_active=True).order_by('name')

    # 각 데이터 소스별로 최신 10개 데이터 가져오기
    sources_with_data = []
    for source in data_sources:
        data_items = CollectedData.objects.filter(
            source=source
        ).order_by('-collected_at')[:10]

        sources_with_data.append({
            'source': source,
            'items': data_items
        })

    context = {
        'subcategory': subcategory,
        'category': subcategory.category,
        'sources_with_data': sources_with_data,
    }

    return render(request, 'core/subcategory_detail.html', context)


def mosquito_test(request):
    """모기 테스트 페이지 (세션 기반 로그인 보호)"""
    if request.session.get('mosquito_auth'):
        is_admin = bool(request.session.get('mosquito_is_admin'))
        return render(request, 'core/mosquito_test.html', {
            'is_admin': is_admin,
            'login_id': request.session.get('mosquito_login_id', ''),
        })

    error = ''
    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = user_store.authenticate(username, password)
        if user:
            request.session['mosquito_auth'] = True
            request.session['mosquito_is_admin'] = user['is_admin']
            request.session['mosquito_login_id'] = user['login_id']
            request.session['mosquito_allowed_devices'] = user['allowed_devices']
            return render(request, 'core/mosquito_test.html', {
                'is_admin': user['is_admin'],
                'login_id': user['login_id'],
            })
        error = '아이디 또는 비밀번호가 올바르지 않습니다.'

    return render(request, 'core/mosquito_login.html', {'error': error})


def beta_view(request):
    """창업시장 베타 페이지 (admin/admin 보호)"""
    if request.session.get('beta_auth'):
        return render(request, 'core/beta_page.html')
    error = ''
    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        if username == 'admin' and password == 'admin':
            request.session['beta_auth'] = True
            return render(request, 'core/beta_page.html')
        error = '아이디 또는 비밀번호가 올바르지 않습니다.'
    return render(request, 'core/beta_login.html', {'error': error})


def beta_logout(request):
    """베타 페이지 로그아웃"""
    if 'beta_auth' in request.session:
        del request.session['beta_auth']
    from django.shortcuts import redirect
    return redirect('/beta/')


def mosquito_logout(request):
    """세션 로그아웃. GET/POST 모두 지원."""
    for k in ('mosquito_auth', 'mosquito_is_admin', 'mosquito_login_id', 'mosquito_allowed_devices'):
        if k in request.session:
            del request.session[k]
    from django.shortcuts import redirect
    return redirect('/mosquito-test/')


def _station_name(raw):
    """관측소 표시명 정제: 코드(SY03서울식물원0043) → 한글 이름(서울식물원).
    앞쪽 영문+숫자 접두 + 뒤쪽 숫자 제거. 정제 결과가 비면 원본 유지.
    """
    s = (raw or '').strip()
    if not s:
        return ''
    # 앞쪽 영문(대소문자) 접두 제거 (예: SY, BD, KH, GH)
    i = 0
    while i < len(s) and (s[i].isascii() and s[i].isalpha()):
        i += 1
    # 권역 약어가 한글로 붙는 경우(GH'서'05...): 한글 1~2자 뒤에 숫자가 오면 그 약어+숫자도 접두로 간주
    if i > 0:
        k = i
        kor = 0
        while k < len(s) and ('가' <= s[k] <= '힣') and kor < 2:
            k += 1; kor += 1
        if kor > 0 and k < len(s) and s[k].isdigit():
            i = k  # 한글 약어까지 접두에 포함
    # 접두의 숫자 제거 (예: SY03, 서05)
    while i < len(s) and s[i].isdigit():
        i += 1
    s2 = s[i:]
    # 뒤쪽 숫자 제거 (예: 0043)
    j = len(s2)
    while j > 0 and s2[j-1].isdigit():
        j -= 1
    s2 = s2[:j].strip()
    return s2 or (raw or '').strip()


def _current_session_user(request):
    """세션에서 복원한 사용자. 없으면 None."""
    if not request.session.get('mosquito_auth'):
        return None
    return {
        'login_id': request.session.get('mosquito_login_id', ''),
        'is_admin': bool(request.session.get('mosquito_is_admin')),
        'allowed_devices': request.session.get('mosquito_allowed_devices'),
    }


def game_notices(request):
    """게임 공지사항 페이지"""
    # 메이플스토리 공지사항
    maple_notices = CollectedData.objects.filter(
        data__game='메이플스토리'
    ).select_related('source').order_by('-collected_at')[:50]

    context = {
        'notices': maple_notices,
        'game_name': '메이플스토리',
    }

    return render(request, 'core/game_notices.html', context)


def _require_mosquito_auth(request):
    """모기 대시보드 세션 인증 체크, 미인증 시 401 JsonResponse 반환"""
    if request.session.get('mosquito_auth'):
        return None
    return JsonResponse({'error': 'Unauthorized'}, status=401)


def _require_admin(request):
    """admin 세션 체크"""
    if request.session.get('mosquito_auth') and request.session.get('mosquito_is_admin'):
        return None
    return JsonResponse({'error': 'Forbidden (admin only)'}, status=403)


@require_GET
def moscom_my_devices(request):
    """현재 로그인 사용자가 접근 가능한 장비 UUID 목록 반환.
    admin: is_admin=True + devices=null(전체 허용)
    user : 허용 UUID 배열
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    su = _current_session_user(request)
    return JsonResponse({
        'login_id': su['login_id'],
        'is_admin': su['is_admin'],
        'allowed_devices': su['allowed_devices'],
    })


def moscom_users_api(request):
    """관리자 전용 사용자 CRUD.
    GET    /users/               → 목록
    POST   /users/               → 추가 (body: login_id, password, allowed_devices[])
    PUT    /users/<login_id>/    → 수정 (body: password? allowed_devices?)
    DELETE /users/<login_id>/    → 삭제
    """
    admin_err = _require_admin(request)
    if admin_err:
        return admin_err
    if request.method == 'GET':
        return JsonResponse({'users': user_store.list_users()})
    # 본문 파싱 (POST/PUT/DELETE)
    try:
        body = _json.loads((request.body or b'').decode('utf-8', errors='replace') or '{}')
    except _json.JSONDecodeError:
        body = {}
    if request.method == 'POST':
        try:
            user_store.create_user(
                login_id=body.get('login_id', ''),
                password=body.get('password', ''),
                allowed_devices=body.get('allowed_devices') or [],
            )
        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)
        return JsonResponse({'ok': True, 'users': user_store.list_users()})
    return JsonResponse({'error': 'method not allowed'}, status=405)


def _build_overview_data(su, date_str='', hour_str=''):
    """종합 현황 탭 + 보고서 재사용용 데이터 빌더.
    허용 장비(su 기준)로만 산출.
    date_str: YYYY-MM-DD 형식, 기준일 변경시 사용. 빈 문자열이면 어제(전일) 기본.
    hour_str: HH (0~23), 지정시 그 시점까지 누적. 빈 문자열이면 하루 전체.
    """
    from datetime import datetime, timedelta, timezone, date as date_cls
    from collections import defaultdict

    # 기준일 결정: date_str 있으면 그 날짜
    # 없으면 영업일 어제(어제 새벽 5시 ~ 오늘 새벽 5시 = 어제 데이터)
    if date_str:
        try:
            target_date = date_cls.fromisoformat(date_str)
        except ValueError:
            target_date = None
    else:
        target_date = None
    if target_date is None:
        try:
            from moscom.timeutil import business_yesterday
            target_date = business_yesterday()
        except Exception:
            kst_now = datetime.now(timezone.utc) + timedelta(hours=9)
            target_date = (kst_now - timedelta(days=1)).date()
    target_iso = target_date.isoformat()
    yday_iso = (target_date - timedelta(days=1)).isoformat()
    # hour_str 처리: 지정시 그 시각까지 cutoff (서버 응답엔 daily 값을 그대로 사용)
    cutoff_hour = None
    if hour_str:
        try:
            cutoff_hour = int(hour_str)
        except ValueError:
            cutoff_hour = None

    # 전역 이상 감지 기준: 51마리 이상이면 이상으로 판정
    ANOMALY_THRESHOLD = 51

    devices = moscom_client.list_devices()
    devices = user_store.filter_devices(su, devices)
    allowed_uuids = {d.get('device_uuid') for d in devices}

    # 우리 DB(moscom.Device)에서 날씨 데이터 + 권역 lookup
    weather_map = {}
    region_map_per_uuid = {}  # uuid → (code, name)
    region_name_by_code = {}
    habitat_map = {}  # uuid → (region_type, form_type) 수동 지정값
    try:
        from moscom.models import Device as MoscomDevice, Region as MoscomRegion
        region_name_by_code = {r.code: r.name for r in MoscomRegion.objects.all()}
        for md in MoscomDevice.objects.filter(device_uuid__in=allowed_uuids):
            weather_map[md.device_uuid] = {
                'temperature': md.temperature,
                'humidity': md.humidity,
                'precipitation': md.precipitation,
                'wind_speed': md.wind_speed,
            }
            region_map_per_uuid[md.device_uuid] = (
                md.region_code or '',
                region_name_by_code.get(md.region_code, md.region_code) or '미지정',
            )
            habitat_map[md.device_uuid] = (md.region_type or '', md.form_type or '')
    except Exception:
        pass

    # 이름/주소 맵
    meta = {}
    for d in devices:
        dv = d.get('device') or {}
        nm = _station_name(dv.get('device_name') or '') or d.get('device_uuid')
        addr = ' '.join(p for p in [dv.get('address_gungu'), dv.get('address_dong')] if p and len(p) < 40 and _valid_kor(p)) or ''
        w = weather_map.get(d.get('device_uuid'), {})
        rc, rn = region_map_per_uuid.get(d.get('device_uuid'), ('', '미지정'))
        rt, ft = habitat_map.get(d.get('device_uuid'), ('', ''))
        meta[d.get('device_uuid')] = {
            'name': nm, 'addr': addr,
            # 51마리 이상 = 이상 (전역 통일)
            'bad_min': ANOMALY_THRESHOLD,
            # 관측소별 실제 임계 (민원 점수 ax1 — 민원가능지역 페이지와 일치시키기 위함)
            'dev_bad_min': ((dv.get('deviceSetting') or {}).get('bad_min')) or 100,
            'region_code': rc,
            'region_name': rn,
            'region_type': rt,
            'form_type': ft,
            'battery': dv.get('battery') or 0,
            'fan': dv.get('fan') or 0,
            'updated_date': dv.get('updated_date'),
            'temperature': w.get('temperature'),
            'humidity': w.get('humidity'),
            'precipitation': w.get('precipitation'),
            'wind_speed': w.get('wind_speed'),
        }

    # 7일 통계 — 전국 전체(필터 전) 보존 후 허용 장비만 필터
    stats_all = moscom_client.get_statistics(device_uuid='', period_type='2', offset=0) or []
    stats = [r for r in stats_all if r.get('device_uuid') in allowed_uuids]

    daily = defaultdict(lambda: defaultdict(int))
    dates_set = set()
    for r in stats:
        u = r.get('device_uuid'); dd = (r.get('created_date') or '')[:10]
        if u and dd:
            daily[u][dd] += (r.get('mosquito_count') or 0)
            dates_set.add(dd)
    sorted_dates = sorted(dates_set)
    # date_str 가 주어진 경우 그 날짜를 기준일로 사용 (없으면 7일 stats 의 마지막 날짜 = 어제)
    today_d = target_iso if target_iso in dates_set else (sorted_dates[-1] if sorted_dates else target_iso)
    yday_d = yday_iso if yday_iso in dates_set else (sorted_dates[-2] if len(sorted_dates) >= 2 else yday_iso)

    # ?hour 지정 시: 그 날짜의 0시 ~ hour시까지의 누적값으로 today_by_dev 재계산
    if cutoff_hour is not None and 0 <= cutoff_hour <= 23:
        try:
            from moscom.models import Collection
            from django.db.models import Sum
            from datetime import datetime as dt
            day_start = dt(target_date.year, target_date.month, target_date.day, 0, 0, 0, tzinfo=timezone.utc) - timedelta(hours=9)
            cutoff_dt = day_start + timedelta(hours=cutoff_hour)
            cutoff_data = (
                Collection.objects
                .filter(device_uuid__in=allowed_uuids,
                        created_date__gte=day_start,
                        created_date__lte=cutoff_dt)
                .values('device_uuid')
                .annotate(total=Sum('mosquito_count'))
            )
            hour_by_dev = {row['device_uuid']: (row['total'] or 0) for row in cutoff_data}
            # daily[today_d] 를 시각별 cutoff 값으로 override
            for u in allowed_uuids:
                daily[u][today_d] = hour_by_dev.get(u, 0)
        except Exception as e:
            logger.warning(f'hour cutoff failed: {e}')

    # 오늘 값
    today_by_dev = {u: daily[u].get(today_d, 0) for u in allowed_uuids} if today_d else {u: 0 for u in allowed_uuids}
    today_total = sum(today_by_dev.values())
    yday_total = sum(daily[u].get(yday_d, 0) for u in allowed_uuids) if yday_d else 0
    change_pct = round((today_total - yday_total) / yday_total * 100) if yday_total else 0

    # 최다 포집 장비
    top_dev = None
    if today_by_dev:
        top_uuid, top_count = max(today_by_dev.items(), key=lambda x: x[1])
        top_dev = {
            'uuid': top_uuid,
            'name': meta.get(top_uuid, {}).get('name', top_uuid),
            'addr': meta.get(top_uuid, {}).get('addr', ''),
            'count': top_count,
        }

    # 장비별 7일 평균 + 추세 + 위험점수 (complaint-risk 축약본)
    warn_count = 0         # 위험 점수 61+ (경고/심각)
    anomaly_today = 0      # 오늘 bad_min 초과
    alert_count = 0        # 경보 발생 — 이상감지 탭과 동일: 기준일 포집 50마리 이상 (관측소당 1회)
    complaint_high = 0     # 민원 위험 61+
    now_utc = datetime.now(timezone.utc)
    offline_count = 0
    check_count = 0
    trust_scores = []
    low_batt_count = 0
    # 장비 이상 판정(요청 기준): 배터리 20% 이하 OR 수신지연 60분 이상
    equip_bad_count = 0      # 위 기준 충족 장비 수
    equip_lowbatt_count = 0  # 그 중 배터리 20% 이하
    equip_delay_count = 0    # 그 중 수신지연 60분 이상

    # 장비별 위험·민원 축약 계산
    # 야간 피크 계산 위해 raw 가져옴 (48h)
    start_iso = (now_utc - timedelta(days=2)).strftime('%Y-%m-%dT%H:%M:%S.000Z')
    end_iso = now_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')
    try:
        raw = moscom_client.get_statistics_by_date(start_dt=start_iso, end_dt=end_iso, aggregation='raw', device_uuid='0')
    except Exception:
        raw = []
    raw = [r for r in (raw or []) if r.get('device_uuid') in allowed_uuids]
    per_dev_hour = defaultdict(lambda: [None] * 24)
    for r in raw:
        u = r.get('device_uuid')
        iso = r.get('created_date') or ''
        try:
            utc_h = int(iso[11:13])
        except Exception:
            continue
        kst_h = (utc_h + 9) % 24
        cnt = r.get('mosquito_count') or 0
        cur = per_dev_hour[u][kst_h]
        if cur is None or cnt > cur:
            per_dev_hour[u][kst_h] = cnt
    dev_night_ratio = {}
    for u, hr in per_dev_hour.items():
        prev = 0; deltas = [0] * 24
        for h in range(24):
            v = hr[h]
            if v is None: continue
            d_ = v - prev; deltas[h] = d_ if d_ > 0 else 0; prev = v
        total = sum(deltas); night = sum(deltas[19:24])
        dev_night_ratio[u] = (night / total) if total > 0 else 0.0

    # 오늘 백분위 계산용
    today_counts_sorted = sorted(today_by_dev.values(), reverse=True)
    n = len(today_counts_sorted)

    def percentile_rank(v):
        if n <= 1: return 100
        above = sum(1 for x in today_counts_sorted if x < v)
        return round((above / (n - 1)) * 100)

    residential_keywords = ['공원', '아파트', '주거', '학교', '초등', '중학', '고등', '어린이집', '유치원']

    def is_resi(m):
        txt = ' '.join([m.get('addr') or '', m.get('name') or ''])
        return any(k in txt for k in residential_keywords)

    # 장비별 상세 계산
    device_rows = []
    for u in allowed_uuids:
        m = meta.get(u, {})
        today_c = today_by_dev.get(u, 0)
        yday_c = daily[u].get(yday_d, 0) if yday_d else 0
        week_vals = [daily[u].get(d, 0) for d in sorted_dates[-7:]]
        week_avg = sum(week_vals) / len(week_vals) if week_vals else 0

        # 위험 점수 4축
        ax1 = min(100, (today_c / m.get('bad_min', 100)) * 100) if m.get('bad_min', 0) > 0 else 0
        ax2 = max(0, min(100, (today_c - yday_c) / yday_c * 100)) if yday_c > 0 else (50 if today_c > 0 else 0)
        if len(week_vals) >= 4:
            half = len(week_vals) // 2
            fh = sum(week_vals[:half]) / half if half > 0 else 0
            lh = sum(week_vals[half:]) / (len(week_vals) - half)
            ax3 = max(0, min(100, ((lh - fh) / fh * 100) / 30 * 100)) if fh > 0 else (50 if lh > 0 else 0)
        else:
            ax3 = 0
        ax4 = percentile_rank(today_c)
        risk_score = round(ax1 * 0.40 + ax2 * 0.25 + ax3 * 0.20 + ax4 * 0.15, 1)
        if risk_score <= 20: risk_lv = '안전'
        elif risk_score <= 40: risk_lv = '관심'
        elif risk_score <= 60: risk_lv = '주의'
        elif risk_score <= 80: risk_lv = '경고'
        else: risk_lv = '심각'
        if risk_score >= 61: warn_count += 1
        if today_c >= m.get('bad_min', 100) and m.get('bad_min', 0) > 0: anomaly_today += 1
        if today_c >= 50: alert_count += 1  # 경보 발생 (이상감지 탭과 동일 기준 50마리, 관측소당 1회)

        # 민원 점수 — 민원가능지역 페이지와 동일하게 관측소별 실제 bad_min 기준
        dev_bm = m.get('dev_bad_min', 100)
        ax_felt = min(100, (today_c / dev_bm) * 100) if dev_bm > 0 else 0
        night_ratio = dev_night_ratio.get(u, 0.0)
        ax_night = min(100, (night_ratio / 0.6) * 100) if night_ratio else 0
        ax_resi = 80 if is_resi(m) else 20
        complaint_score = round(ax_felt * 0.35 + ax2 * 0.25 + ax_night * 0.25 + ax_resi * 0.15, 1)
        if complaint_score <= 30: cp_lv = '낮음'
        elif complaint_score <= 60: cp_lv = '보통'
        else: cp_lv = '높음'
        if cp_lv == '높음': complaint_high += 1  # 테이블 등급과 동일 기준

        # 장비 신뢰도 (equipment-health 축약)
        battery = m.get('battery', 0)
        fan = m.get('fan', 0)
        if battery < 15: low_batt_count += 1
        ax_bat = 100 if battery >= 50 else 80 if battery >= 30 else 60 if battery >= 20 else 35 if battery >= 10 else 10
        ax_fan = 100 if fan == 1 else 30
        try:
            ud = m.get('updated_date') or ''
            udt = datetime.fromisoformat(ud.replace('Z', '+00:00'))
            delay_min = (now_utc - udt).total_seconds() / 60
        except Exception:
            delay_min = 99999
        if delay_min <= 120: ax_sig = 100
        elif delay_min <= 360: ax_sig = 80
        elif delay_min <= 720: ax_sig = 55
        elif delay_min <= 1440: ax_sig = 30
        else: ax_sig = 5
        zero_total_3 = sum(week_vals[-3:]) if week_vals else 0
        ax_zero = 40 if (len(week_vals) >= 3 and zero_total_3 == 0) else (65 if zero_total_3 == 0 else 100)
        trust_score = round(ax_bat * 0.25 + ax_fan * 0.20 + ax_sig * 0.35 + ax_zero * 0.20, 1)
        trust_scores.append(trust_score)

        if ax_sig <= 30: equip_status = '오프라인'; offline_count += 1
        elif trust_score < 60 or battery < 15 or (fan == 0 and ax_sig >= 55): equip_status = '점검필요'; check_count += 1
        else: equip_status = '정상'

        # 장비 이상 판정(요청 기준): 배터리 20% 이하만
        is_lowbatt = battery <= 20
        is_delayed = delay_min >= 60
        if is_lowbatt:
            equip_lowbatt_count += 1
        if is_delayed:
            equip_delay_count += 1
        if is_lowbatt:
            equip_bad_count += 1

        # 추세 방향: 기준일(today_c) vs 전일(yday_c) 비교
        diff = today_c - yday_c
        if diff > 3:
            trend_dir = '↑'
        elif diff < -3:
            trend_dir = '↓'
        else:
            trend_dir = '→'

        device_rows.append({
            'uuid': u, 'name': m.get('name'), 'addr': m.get('addr'),
            'today': today_c, 'yday': yday_c, 'week_avg': round(week_avg, 1),
            'risk_score': risk_score, 'risk_level': risk_lv,
            'complaint_score': complaint_score, 'complaint_level': cp_lv,
            'trust_score': trust_score, 'status': equip_status,
            'battery': battery, 'trend': trend_dir,
            'bad_min': m.get('bad_min', 100),
            # 날씨 (Open-Meteo)
            'temperature': m.get('temperature'),
            'humidity': m.get('humidity'),
            'precipitation': m.get('precipitation'),
            'wind_speed': m.get('wind_speed'),
            'updated_date': m.get('updated_date'),
            # 권역
            'region_code': m.get('region_code') or '',
            'region_name': m.get('region_name') or '미지정',
            # 2축 분류 (수동 지정값 — 비면 JS에서 자동추정)
            'region_type': m.get('region_type') or '',
            'form_type': m.get('form_type') or '',
        })

    # 포집량 내림차순 정렬
    device_rows.sort(key=lambda r: r['today'], reverse=True)

    avg_trust = round(sum(trust_scores) / len(trust_scores), 1) if trust_scores else 0

    # ── 전국 비교 지표 (필터 전 전체 stats 기준, today_d 하루) ──
    # admin: 전국 평균 + 전국 최고 관측소명/마리수
    # 비admin: 전국 평균 + 전국 최고 마리수(관측소명 숨김) + 해당 지역(권역) 전체 평균
    national = None
    try:
        nat_today = defaultdict(int)  # uuid → today_d 포집 합
        for r in stats_all:
            if (r.get('created_date') or '')[:10] == today_d:
                u = r.get('device_uuid')
                if u:
                    nat_today[u] += (r.get('mosquito_count') or 0)
        if nat_today:
            nat_vals = list(nat_today.values())
            nat_avg = round(sum(nat_vals) / len(nat_vals), 1)
            top_uuid_nat = max(nat_today, key=lambda x: nat_today[x])
            top_cnt_nat = nat_today[top_uuid_nat]
            is_admin = bool(su.get('is_admin')) if su else False
            # 전국 device명 매핑 (admin 최고 관측소명 표시용)
            top_name_nat = top_uuid_nat
            try:
                from moscom.models import Device as MD
                md = MD.objects.filter(device_uuid=top_uuid_nat).first()
                if md:
                    top_name_nat = _station_name(md.device_name) or top_uuid_nat
            except Exception:
                pass
            national = {
                'count': len(nat_vals),
                'avg': nat_avg,
                'top_count': top_cnt_nat,
                'is_admin': is_admin,
            }
            if is_admin:
                national['top_name'] = top_name_nat
            else:
                # 사용자 권역(허용 장비들의 region_code) 전체 평균
                my_codes = {meta.get(u, {}).get('region_code') for u in allowed_uuids}
                my_codes.discard('')
                region_avg = None
                if my_codes:
                    region_uuids_nat = set()
                    try:
                        from moscom.models import Device as MD2
                        region_uuids_nat = set(
                            MD2.objects.filter(region_code__in=my_codes, is_active=True)
                            .values_list('device_uuid', flat=True)
                        )
                    except Exception:
                        pass
                    rvals = {u: nat_today[u] for u in region_uuids_nat if u in nat_today}
                    if rvals:
                        region_avg = round(sum(rvals.values()) / len(rvals), 1)
                        # 지역 최고 관측소 + 마리수
                        rtop_uuid = max(rvals, key=lambda x: rvals[x])
                        national['region_top_count'] = rvals[rtop_uuid]
                        national['region_top_name'] = meta.get(rtop_uuid, {}).get('name') or rtop_uuid
                national['region_avg'] = region_avg
                # 종합현황 전용 표시명(overview_name) 우선, 없으면 일반 표시명/코드
                reg_labels = set()
                try:
                    from moscom.models import Region as RG
                    rmap = {r.code: (r.overview_name or r.name or r.code) for r in RG.objects.all()}
                except Exception:
                    rmap = {}
                for c in my_codes:
                    reg_labels.add(rmap.get(c, c))
                national['region_names'] = sorted(reg_labels - {'', None})
    except Exception as e:
        logger.warning(f'national compare failed: {e}')

    return {
        'today': today_d,
        'yday': yday_d,
        'kpi': {
            'today_total': today_total,
            'yday_total': yday_total,
            'change_pct': change_pct,
            'top_dev': top_dev,
            'warn_count': warn_count,
            'anomaly_today': anomaly_today,
            'alert_count': alert_count,
            'complaint_high': complaint_high,
            'equip_bad': equip_bad_count,
            'equip_lowbatt': equip_lowbatt_count,
            'equip_delay': equip_delay_count,
            'offline_count': offline_count,
            'check_count': check_count,
            'low_batt_count': low_batt_count,
            'avg_trust': avg_trust,
            'total_devices': len(allowed_uuids),
        },
        'national': national,
        'devices': device_rows,
    }


@require_GET
def moscom_overview(request):
    """종합 현황 탭용 집계 API. 허용 장비 기준.
    ?date=YYYY-MM-DD&hour=HH 옵션 — 지정시 그 시점 기준 (없으면 어제 = 기본값).
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    su = _current_session_user(request)
    try:
        date_str = (request.GET.get('date') or '').strip()
        hour_str = (request.GET.get('hour') or '').strip()
        data = _build_overview_data(su, date_str=date_str, hour_str=hour_str)
        return JsonResponse(data)
    except Exception as e:
        logger.exception('overview failed')
        return JsonResponse({'error': str(e)}, status=500)


def _build_report_body(period, base_date, su, request):
    """보고서 본문(GPT) + 요약 데이터 구성. (report_text, summary, scoped_uuids, source, payload_summary)"""
    from datetime import datetime, timedelta, timezone
    from collections import defaultdict
    from django.conf import settings as dj_settings

    # 기준일 규칙: 선택일 당일은 수집 진행 중이라 부적합 → 항상 "선택일의 전날"을 데이터 기준일로.
    #  - 미지정 시: 업무일 기준 어제
    #  - 지정 시(예: 7/8 선택): 그 전날(7/7)로 시프트
    try:
        from moscom.timeutil import business_yesterday
        default_base = business_yesterday()
    except Exception:
        default_base = datetime.now(timezone(timedelta(hours=9))).date() - timedelta(days=1)
    if not base_date:
        base_date = default_base.isoformat()
    else:
        try:
            sel = datetime.strptime(base_date[:10], '%Y-%m-%d').date()
            base_date = (sel - timedelta(days=1)).isoformat()   # 선택일의 전날을 기준일로
        except Exception:
            base_date = default_base.isoformat()

    start_s, end_s = report_store.period_range(period, base_date)

    devices = moscom_client.list_devices()
    devices = user_store.filter_devices(su, devices)
    allowed_uuids = [d.get('device_uuid') for d in devices]

    # 일별 집계 (기간 범위)
    start_iso = datetime.strptime(start_s, '%Y-%m-%d').strftime('%Y-%m-%dT00:00:00.000Z')
    end_iso = (datetime.strptime(end_s, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%dT00:00:00.000Z')
    daily_records = moscom_client.get_statistics_by_date(
        start_dt=start_iso, end_dt=end_iso, aggregation='day', device_uuid='0'
    )
    allowed_set = set(allowed_uuids)
    daily_records = [r for r in (daily_records or []) if r.get('device_uuid') in allowed_set]

    daily = defaultdict(lambda: defaultdict(int))
    for r in daily_records:
        u = r.get('device_uuid')
        d = (r.get('created_date') or '')[:10]
        if u and d:
            daily[u][d] += (r.get('mosquito_count') or 0)

    # 날짜 배열
    dates_in_range = sorted({(r.get('created_date') or '')[:10] for r in daily_records if r.get('created_date')})
    total_in_period = sum(sum(v.values()) for v in daily.values())
    # 일평균
    n_days = max(1, len(dates_in_range))
    avg_per_day = round(total_in_period / n_days, 1)

    # 51마리 전역 임계값 (행정 판단과 동일)
    ANOMALY_THRESHOLD = 51

    # moscom DB lookup (권역/기상)
    try:
        from moscom.models import Device as MoscomDevice, Region as MoscomRegion
        region_name_by_code = {r.code: r.name for r in MoscomRegion.objects.all()}
        uuid_to_md = {md.device_uuid: md for md in MoscomDevice.objects.all()}
    except Exception:
        region_name_by_code, uuid_to_md = {}, {}

    # 장비 이름 매핑 (권역명 포함)
    name_map = {}
    for d in devices:
        dv = d.get('device') or {}
        nm = _station_name(dv.get('device_name') or '') or d.get('device_uuid')
        addr = ' '.join(p for p in [dv.get('address_gungu'), dv.get('address_dong')] if p and len(p) < 40) or ''
        md = uuid_to_md.get(d.get('device_uuid'))
        rc = (md.region_code if md else '') or ''
        rname = region_name_by_code.get(rc, rc) or '미지정'
        name_map[d.get('device_uuid')] = {
            'name': nm, 'addr': addr,
            'bad_min': ANOMALY_THRESHOLD,
            'region_name': rname,
            'region_code': rc,
            # 사이트 예측(moscom_predict)과 동일한 입력을 위해 sido·weather 포함
            'sido': (md.address_sido if md else (dv.get('address_sido') or '')) or '',
            'weather': {
                'temperature': md.temperature if md else None,
                'humidity': md.humidity if md else None,
                'precipitation': md.precipitation if md else None,
                'wind_speed': md.wind_speed if md else None,
            } if md else {},
        }

    # 장비별 기간 합계 Top 5
    dev_totals = sorted(
        [(u, sum(daily[u].values())) for u in daily.keys()],
        key=lambda x: x[1], reverse=True
    )[:5]

    # 권역별 기간 합계
    region_totals = defaultdict(lambda: {'total': 0, 'devices': 0, 'top_dev': '', 'top_count': 0})
    for u in allowed_uuids:
        v = sum(daily[u].values())
        meta = name_map.get(u, {})
        rname = meta.get('region_name', '미지정')
        region_totals[rname]['total'] += v
        region_totals[rname]['devices'] += 1
        if v > region_totals[rname]['top_count']:
            region_totals[rname]['top_count'] = v
            region_totals[rname]['top_dev'] = meta.get('name', '')
    region_sorted = sorted(region_totals.items(), key=lambda kv: -kv[1]['total'])

    # 이상 감지 (51마리 이상)
    anomalies = []
    for u in allowed_uuids:
        bm = ANOMALY_THRESHOLD
        for dd, c in daily[u].items():
            if c >= bm:
                anomalies.append({
                    'name': name_map.get(u, {}).get('name', u),
                    'addr': name_map.get(u, {}).get('addr', ''),
                    'region_name': name_map.get(u, {}).get('region_name', '미지정'),
                    'date': dd, 'count': c, 'bad_min': bm,
                })
    anomalies.sort(key=lambda a: a['count'], reverse=True)
    anomalies = anomalies[:15]

    # 방역 계획 (해당 기간 겹치는 것)
    visible = allowed_set if not su.get('is_admin') else None
    all_plans = remedy_store.list_plans(visible_uuids=visible)
    method_map = {m['key']: m for m in remedy_store.list_methods()}
    plans_in_range = []
    plans_executed_count = 0
    for p in all_plans:
        sd = p.get('scheduled_date') or ''
        m = method_map.get(p['method_key'], {})
        in_range = (start_s <= sd <= end_s)
        if in_range or sd >= start_s:
            plans_in_range.append({
                'device': name_map.get(p['device_uuid'], {}).get('name') or p['device_uuid'],
                'method': m.get('name', p['method_key']),
                'scheduled_date': sd,
                'reduction_pct': m.get('reduction_pct'),
                'executed': in_range,
                'worker': p.get('worker') or '',
                'volume_l': p.get('volume_l'),
            })
            if in_range:
                plans_executed_count += 1

    # 장비 상태 — 30% 미만 카운트 (행정 판단과 동일)
    now_utc = datetime.now(timezone.utc)
    offline_count = 0; low_batt_count = 0
    batt_vals = []
    for d in devices:
        dv = d.get('device') or {}
        b = dv.get('battery')
        if b is not None:
            batt_vals.append(b)
        if (b or 100) < 30:
            low_batt_count += 1
        ud = dv.get('updated_date') or ''
        try:
            udt = datetime.fromisoformat(ud.replace('Z', '+00:00'))
            if (now_utc - udt).total_seconds() / 60 > 1440:
                offline_count += 1
        except Exception:
            offline_count += 1
    avg_battery = round(sum(batt_vals) / len(batt_vals)) if batt_vals else 0

    # 기상 평균 + 매개체 위험 평가
    temps = [m.temperature for m in uuid_to_md.values() if m.device_uuid in allowed_set and m.temperature is not None]
    humids = [m.humidity for m in uuid_to_md.values() if m.device_uuid in allowed_set and m.humidity is not None]
    precs = [m.precipitation for m in uuid_to_md.values() if m.device_uuid in allowed_set and m.precipitation is not None]
    avg_temp = round(sum(temps)/len(temps), 1) if temps else None
    avg_humid = round(sum(humids)/len(humids), 1) if humids else None
    avg_precip = round(sum(precs)/len(precs), 2) if precs else 0

    vector_risks = []
    temp_zone = humid_zone = None
    try:
        from moscom.health_knowledge import assess_vector_risks, thermal_zone_for, humidity_zone_for
        vector_risks = assess_vector_risks(avg_temp, avg_humid, avg_precip)
        temp_zone = thermal_zone_for(avg_temp)
        humid_zone = humidity_zone_for(avg_humid)
    except Exception as e:
        logger.warning(f'health_knowledge failed: {e}')

    # 기간 추세 — 전반 vs 후반
    sorted_dates = sorted(dates_in_range)
    trend_text = ''
    if len(sorted_dates) >= 4:
        half = len(sorted_dates) // 2
        first_half = sum(sum(daily[u].get(dd, 0) for u in allowed_uuids) for dd in sorted_dates[:half])
        second_half = sum(sum(daily[u].get(dd, 0) for u in allowed_uuids) for dd in sorted_dates[half:])
        fh_avg = first_half / max(1, half)
        sh_avg = second_half / max(1, len(sorted_dates) - half)
        if fh_avg > 0:
            tr_pct = round((sh_avg - fh_avg) / fh_avg * 100)
            if tr_pct > 15:
                trend_text = f'기간 추세: 전반 일평균 {round(fh_avg)} → 후반 {round(sh_avg)} (+{tr_pct}%, 상승)'
            elif tr_pct < -15:
                trend_text = f'기간 추세: 전반 일평균 {round(fh_avg)} → 후반 {round(sh_avg)} ({tr_pct}%, 하강)'
            else:
                trend_text = f'기간 추세: 전반 {round(fh_avg)} ↔ 후반 {round(sh_avg)} (안정)'

    # 7일 시계열 (전체 합)
    last7_dates = sorted_dates[-7:] if len(sorted_dates) >= 7 else sorted_dates
    last7_timeseries = []
    for dd in last7_dates:
        t = sum(daily[u].get(dd, 0) for u in allowed_uuids)
        last7_timeseries.append({'date': dd, 'total': t})

    # 최근 14일 방역 이력 (참고용)
    try:
        from datetime import date as _date
        cutoff_dt = _date.today() - timedelta(days=14)
        cutoff = cutoff_dt.strftime('%Y-%m-%d')
    except Exception:
        cutoff = ''
    recent_plans = []
    for p in all_plans:
        sd = p.get('scheduled_date') or ''
        if cutoff and sd < cutoff:
            continue
        m = method_map.get(p['method_key'], {})
        recent_plans.append({
            'device': name_map.get(p['device_uuid'], {}).get('name') or p['device_uuid'],
            'method': m.get('name', p['method_key']),
            'scheduled_date': sd,
            'reduction_pct': m.get('reduction_pct'),
        })
    recent_plans.sort(key=lambda x: x['scheduled_date'], reverse=True)
    recent_plans = recent_plans[:15]

    # GPT 프롬프트
    period_label = {'daily': '일간', 'weekly': '주간', 'monthly': '월간'}.get(period, '일간')
    top_devs_text = '\n'.join(
        f"  - {name_map.get(u, {}).get('name')} [{name_map.get(u, {}).get('region_name', '미지정')}] ({name_map.get(u, {}).get('addr') or '주소미상'}): 기간 합계 {v}마리"
        for u, v in dev_totals
    ) or '  - 데이터 없음'
    anomaly_text = '\n'.join(
        f"  - {a['date']} {a['name']} [{a.get('region_name','미지정')}] ({a['addr'] or '주소미상'}) — {a['count']}마리 / 임계 {a['bad_min']}마리"
        for a in anomalies
    ) or '  - 임계(51마리) 초과 이상 감지 없음'
    plans_text = '\n'.join(
        f"  - {p['device']} / {p['method']} / 예정 {p['scheduled_date']} / 감소율 {p['reduction_pct']}% {'[기간내]' if p.get('executed') else ''}"
        for p in plans_in_range[:10]
    ) or '  - 해당 기간 방역 계획 없음'

    # 권역별 합계 Top 10
    region_text = '\n'.join(
        f"  - {rname}: 합계 {rd['total']}마리 · 장비 {rd['devices']}대 · 최다 {rd['top_dev'] or '-'} ({rd['top_count']}마리)"
        for rname, rd in region_sorted[:10]
    ) or '  - 권역 정보 없음'

    # 7일 시계열 텍스트
    last7_text = '\n'.join(
        f"  - {ts['date']}: {ts['total']}마리"
        for ts in last7_timeseries
    ) or '  - 시계열 데이터 부족'

    # 기상-생리학 텍스트
    thermal_text = (
        f"  - 평균기온 {avg_temp}°C → {temp_zone.get('effect', '-')} "
        f"(유충 발달 {temp_zone.get('larval_days','-')}일, 성충 수명 {temp_zone.get('adult_lifespan','-')})"
        if (avg_temp is not None and temp_zone) else '  - 기온 데이터 없음'
    )
    humid_text = (
        f"  - 평균습도 {avg_humid}% → {humid_zone.get('effect', '-')}"
        if (avg_humid is not None and humid_zone) else '  - 습도 데이터 없음'
    )
    precip_text = f"  - 평균 강수 {avg_precip}mm/일 (* 강수 후 5~7일 1차 우화 피크)" if avg_precip else '  - 유의미 강수 없음'

    # 매개체 위험 평가 텍스트
    if vector_risks:
        vector_text = '\n'.join(
            f"  - {vr['species']} ({', '.join(vr['diseases'])}) — 위험 {vr['risk_level']} ({vr['risk_score']}) · 활동 {vr['active_hours']} · 서식 {vr['breeding_site']}"
            for vr in vector_risks[:4]
        )
    else:
        vector_text = '  - 기상 데이터 부족으로 평가 보류'

    # 최근 방역 이력 텍스트
    recent_plans_text = '\n'.join(
        f"  - {p['scheduled_date']} {p['device']} / {p['method']} (-{p['reduction_pct']}%)"
        for p in recent_plans[:10]
    ) or '  - 최근 14일간 시행/예정 방역 없음'

    # 종합 현황 KPI + 전국 지표 (최신일 기준) — 보고서 AI 요약 근거. 본문 "핵심 수치"와 중복되지 않도록 입력 데이터로만 활용.
    try:
        ov = _build_overview_data(su)
        k = ov.get('kpi') or {}
        nat = ov.get('national') or {}
        # 전국 지표 라인 구성 (admin 여부에 따라)
        nat_lines = ''
        if nat:
            nat_lines = (
                f"  - 전국 관측소 수: {nat.get('count', 0)}개 · 전국 일평균: {nat.get('avg', 0)}마리\n"
            )
            if nat.get('is_admin'):
                if nat.get('top_name'):
                    nat_lines += f"  - 전국 최다 관측소: {nat.get('top_name')} · {nat.get('top_count', 0)}마리\n"
            else:
                if nat.get('region_avg') is not None:
                    region_label = ' · '.join(nat.get('region_names') or []) or '담당 권역'
                    nat_lines += f"  - {region_label} 평균: {nat.get('region_avg')}마리 (전국 대비 {round((nat.get('region_avg') or 0) / nat.get('avg') , 2) if nat.get('avg') else '-'}배)\n"
                    if nat.get('region_top_name'):
                        nat_lines += f"  - 담당 권역 최다 관측소: {nat.get('region_top_name')} · {nat.get('region_top_count', 0)}마리\n"
        overview_block = (
            "\n■ 종합 현황·전국 지표 (보고일 기준, 입력 데이터)\n"
            f"  - 금일 포집 합계: {k.get('today_total', 0)}마리 (전일 대비 {k.get('change_pct', 0):+d}%)\n"
            f"  - 최다 포집 관측소: {((k.get('top_dev') or {}).get('name') or '-')} · {(k.get('top_dev') or {}).get('count', 0)}마리\n"
            f"  - 경고 이상 관측소: {k.get('warn_count', 0)}개 · 경보 발생: {k.get('alert_count', 0)}곳\n"
            f"  - 민원 가능 '높음': {k.get('complaint_high', 0)}개\n"
            f"  - 장비 이상: {k.get('equip_bad', 0)}대 (배터리 20%↓ {k.get('equip_lowbatt', 0)}/수신지연 1h↑ {k.get('equip_delay', 0)})\n"
            f"  - 평균 데이터 신뢰도: {k.get('avg_trust', 0)}%\n"
            f"{nat_lines}"
        )
    except Exception as _e:
        overview_block = ''
        k = {}

    payload_summary = f"""[{period_label} 모기 발생 감시 보고 · {start_s} ~ {end_s} ({n_days}일)]
{overview_block}
■ 기간 전체 수치 (이상감지 임계 51마리)
  - 전체 장비 수: {len(devices)}대
  - 기간 합계 포집량: {total_in_period}마리 (일평균 {avg_per_day}마리)
  - 측정 일수: {n_days}일
  - 오프라인 장비: {offline_count}대 (24시간 이상 미수신)
  - 배터리 30% 미만: {low_batt_count}대 / 평균 배터리: {avg_battery}%
  - 방역 시행/예정: {len(plans_in_range)}건 (기간 내 시행 {plans_executed_count}건)
  - {trend_text or '기간 추세: 산출 불가'}

■ 권역별 기간 합계 (Top 10)
{region_text}

■ 기간 상위 포집 장비 (Top 5)
{top_devs_text}

■ 7일 시계열 (전체 합)
{last7_text}

■ 포집량 이상 감지 (≥51마리 일별 기준)
{anomaly_text}

■ 기상-모기 생리학 분석
{thermal_text}
{humid_text}
{precip_text}

■ 매개체별 위험 평가 (기상 기반 적합도)
{vector_text}

■ 해당 기간 방역 내역
{plans_text}

■ 최근 14일 방역 이력 (참고)
{recent_plans_text}
"""

    api_key = getattr(dj_settings, 'OPENAI_API_KEY', '') or ''
    if not api_key:
        report_text = '※ OpenAI API 키가 서버에 설정되지 않아 AI 분석이 불가합니다.\n\n' + payload_summary
        source = 'fallback'
    else:
        try:
            import openai
            client = openai.OpenAI(api_key=api_key)
            period_instruction = {
                'daily':   '오늘 하루의 발생/대응 현황을 중심으로 요약하십시오.',
                'weekly':  '최근 7일간의 추세·권역별 분포·매개체 위험을 중심으로 종합하십시오.',
                'monthly': '최근 30일간의 장기 추세, 계절성, 권역간 격차, 매개체 활동 특성을 반영하여 종합하십시오.',
            }.get(period, '요약하십시오.')
            resp = client.chat.completions.create(
                model='gpt-4o-mini',
                messages=[
                    {'role': 'system', 'content': (
                        f'당신은 지자체 보건소 감염병관리팀의 {period_label} 모기 발생 감시 보고서 작성을 보조하는 '
                        '보건학(매개체 감시·방역) 전문가 AI입니다. '
                        f'{period_instruction} '
                        '주어진 실측 데이터, 기상-모기 생리학 분석, 매개체 위험 평가를 근거로 결재용 공문 형식의 한국어 보고서를 작성하십시오. '
                        '다음 8개 섹션을 반드시 포함하십시오. "핵심 수치 요약"은 반드시 1) 섹션 하나로만 작성하고, '
                        '입력의 "종합 현황·전국 지표" 내용은 별도 섹션으로 분리하지 말고 1) 섹션 안에 녹여 작성하십시오(같은 내용 중복 금지):\n'
                        '1) 핵심 수치 요약 (포집량/장비/이상감지/방역 + 전국 지표·전국 대비 비교 포함)\n'
                        '2) 권역별 발생 양상 (Top 권역, 격차, 우점 권역)\n'
                        '3) 시계열·추세 분석 (전반·후반 비교, 7일 추이)\n'
                        '4) 이상 발생 및 주요 관측 (51마리 이상 일자/장비, 패턴)\n'
                        '5) 기상-매개체 위험 평가 (기온·습도·강수 조건과 매개종 활동, 우화 주기)\n'
                        '6) 방역 실시 현황 및 효과 (시행 건수, 방법별 분포)\n'
                        '7) 향후 대응 권고 (구체 방역 방법·시기·대상 장비 명시, 매개종-방역 매칭)\n'
                        '8) 참고·특이사항 (장비 상태, 데이터 신뢰도, 후속 모니터링).\n\n'
                        '규칙:\n'
                        '- 각 섹션은 "■" 기호로 시작, 불릿은 "- "로 시작합니다.\n'
                        '- 공공기관 행정 어조(…함, …필요함, …권고함, …요망)를 사용하십시오.\n'
                        '- 모든 수치·장비명·권역명·날짜는 제공된 데이터에서만 인용하고 임의로 생성하지 마십시오.\n'
                        '- 51마리 이상 일별 포집은 "이상감지"로 명시하고 해당 장비·권역을 구체적으로 적시하십시오.\n'
                        '- 기상-생리학 섹션은 우화 기간/성충 수명/매개종 활동시간을 인용하여 과학적 근거를 제시하십시오.\n'
                        '- 방역 권고는 Bti·잔류분무·ULV·용기제거 중 데이터 상황에 맞는 방법을 명시하십시오.\n'
                        '- 결재란·서명란·작성자/날짜 줄은 포함하지 마십시오 (화면에서 별도 렌더링됩니다).\n'
                        '- 단순 나열보다 보건학적 인과와 행정 우선순위가 드러나도록 작성하십시오.'
                    )},
                    {'role': 'user', 'content': payload_summary},
                ],
                temperature=0.4,
                max_tokens=1800,
            )
            report_text = resp.choices[0].message.content.strip()
            source = 'openai:gpt-4o-mini'
        except Exception as e:
            logger.exception('OpenAI call failed in report')
            report_text = f'※ AI 분석 중 오류: {e}\n\n' + payload_summary
            source = 'error'

    # 섹션 2: 관측소별 현황 (ov['devices']에 이미 계산되어 있음)
    stations_section = []
    if ov:
        for d in ov.get('devices', []):
            stations_section.append({
                'name': d['name'], 'addr': d.get('addr', ''),
                'risk_level': d['risk_level'], 'risk_score': d['risk_score'],
                'today': d['today'], 'yday': d.get('yday', 0), 'week_avg': d.get('week_avg', 0),
                'complaint_level': d['complaint_level'],
                'status': d['status'], 'trust_score': d['trust_score'],
                'trend': d.get('trend', '·'),
            })

    # 섹션 3: 일간이면 시간별 히트맵 (48h raw로 허용 장비의 시간×장비)
    heatmap_section = None
    if period == 'daily':
        try:
            now_for_hm = datetime.now(timezone.utc)
            hm_start = (now_for_hm - timedelta(days=2)).strftime('%Y-%m-%dT%H:%M:%S.000Z')
            hm_end = now_for_hm.strftime('%Y-%m-%dT%H:%M:%S.000Z')
            hm_raw = moscom_client.get_statistics_by_date(
                start_dt=hm_start, end_dt=hm_end, aggregation='raw', device_uuid='0'
            )
            hm_raw = [r for r in (hm_raw or []) if r.get('device_uuid') in allowed_set]
            per_dev_hour = defaultdict(lambda: [None] * 24)
            for r in hm_raw:
                u = r.get('device_uuid')
                iso = r.get('created_date') or ''
                try:
                    utc_h = int(iso[11:13])
                except Exception:
                    continue
                kst_h = (utc_h + 9) % 24
                cnt = r.get('mosquito_count') or 0
                cur = per_dev_hour[u][kst_h]
                if cur is None or cnt > cur:
                    per_dev_hour[u][kst_h] = cnt
            # 시간별 히트맵 — 시간별 히트맵 탭(2p)과 동일하게 수집창 18:00~익일 05:00 (12칸)만
            HM_HOURS = [18, 19, 20, 21, 22, 23, 0, 1, 2, 3, 4, 5]
            rows = []
            for u in allowed_uuids:
                hr = per_dev_hour.get(u, [None]*24)
                deltas24 = [0]*24; prev = 0
                for h in range(24):
                    v = hr[h]
                    if v is None: continue
                    d_ = v - prev; deltas24[h] = d_ if d_ > 0 else 0; prev = v
                deltas = [deltas24[h] for h in HM_HOURS]  # 수집창 순서로 재배열
                total = sum(deltas)
                rows.append({
                    'name': name_map.get(u, {}).get('name') or u,
                    'hours': deltas, 'total': total,
                })
            rows.sort(key=lambda x: x['total'], reverse=True)
            heatmap_section = rows[:30]  # 상위 30대
        except Exception:
            heatmap_section = None

    # 섹션 4: 이상 감지 상세 (anomalies는 이미 있음)
    anomaly_section = [
        {
            'name': a['name'], 'addr': a['addr'], 'date': a['date'],
            'region_name': a.get('region_name', '미지정'),
            'count': a['count'], 'bad_min': a['bad_min'],
            'pct_over': round((a['count'] - a['bad_min']) / a['bad_min'] * 100) if a['bad_min'] else 0,
        }
        for a in anomalies
    ]

    # 섹션 5: AI 예측 — 사이트(moscom_predict)와 동일하게: 기준일 직전 10일 실측 + 기상 입력
    predict_section = None
    try:
        p_days = 3 if period == 'daily' else 7 if period == 'weekly' else 14
        # 예측용 history: 기준일(end_s) 직전 10일을 별도 조회 (lag7까지 확보, 사이트와 동일 조건)
        pred_hist = defaultdict(list)
        try:
            _end_d = datetime.strptime(end_s, '%Y-%m-%d').date()
            _start_d = _end_d - timedelta(days=10)
            _ph_start = _start_d.strftime('%Y-%m-%dT00:00:00.000Z')
            _ph_end = (_end_d + timedelta(days=1)).strftime('%Y-%m-%dT00:00:00.000Z')
            _ph_records = moscom_client.get_statistics_by_date(
                start_dt=_ph_start, end_dt=_ph_end, aggregation='day', device_uuid='0'
            )
            _ph_daily = defaultdict(lambda: defaultdict(int))
            for r in (_ph_records or []):
                u = r.get('device_uuid'); dd = (r.get('created_date') or '')[:10]
                if u in allowed_set and dd and dd <= end_s:   # 기준일까지만(그 이후 제외)
                    _ph_daily[u][dd] += (r.get('mosquito_count') or 0)
            for u in name_map:
                for dd in sorted(_ph_daily[u].keys()):
                    pred_hist[u].append({'date': dd, 'count': _ph_daily[u][dd]})
        except Exception:
            logger.exception('predict history fetch failed; falling back to period daily')
            for u in name_map:
                for dd in sorted(daily[u].keys()):
                    pred_hist[u].append({'date': dd, 'count': daily[u].get(dd, 0)})

        p_inputs = []
        for u, m in name_map.items():
            p_inputs.append({
                'uuid': u, 'name': m['name'], 'region': m.get('addr') or '',
                'history': pred_hist.get(u, []),
                # 사이트 예측과 동일한 입력 (기상·권역 반영)
                'region_code': m.get('region_code') or '',
                'sido': m.get('sido') or '',
                'weather': m.get('weather') or {},
            })
        raw_preds = predictor.predict_for_devices(p_inputs, days_ahead=p_days)
        # 방역 효과 적용 — 모기지수 보존
        def _grade_count(n):
            if n <= 10: return '안전'
            if n <= 50: return '관심'
            if n <= 100: return '주의'
            if n <= 200: return '경고'
            return '위험'
        def _grade_idx(v):
            if v is None: return None
            if v < 25: return '쾌적'
            if v < 50: return '관심'
            if v < 75: return '주의'
            return '불쾌'
        for p in raw_preds:
            new_preds = []
            for pp in p['predictions']:
                factor, _ = remedy_store.adjustment_factor(p['uuid'], pp.get('date'))
                orig = pp.get('predicted') or 0
                orig_idx = pp.get('predicted_index')
                adj_idx = round(max(0.0, min(100.0, orig_idx * factor)), 1) if orig_idx is not None else None
                new_preds.append({
                    'date': pp['date'],
                    'predicted': int(round(orig * factor)),
                    'predicted_raw': orig,
                    'predicted_index': adj_idx,
                    'predicted_index_raw': round(orig_idx, 1) if orig_idx is not None else None,
                    'grade': _grade_idx(adj_idx),
                    'remedy_factor': round(factor, 3),
                })
            p['predictions'] = new_preds
            ps = [x['predicted'] for x in new_preds]
            idxs = [x['predicted_index'] for x in new_preds if x['predicted_index'] is not None]
            p['max_predicted'] = max(ps) if ps else 0
            if idxs:
                avg_idx = sum(idxs) / len(idxs)
                p['max_index'] = round(max(idxs), 1)
                p['avg_index'] = round(avg_idx, 1)
                p['grade'] = _grade_idx(avg_idx)
            else:
                p['max_index'] = None
                p['avg_index'] = None
                p['grade'] = _grade_count(p['max_predicted'])
        raw_preds.sort(key=lambda x: x.get('max_predicted', 0), reverse=True)
        # 위험 점수 상위 10대 + 날짜별 합계
        predict_section = {
            'devices': [
                {
                    'name': p['name'], 'grade': p['grade'],
                    'max_predicted': p['max_predicted'],
                    'max_index': p.get('max_index'),
                    'avg_index': p.get('avg_index'),
                    'predictions': p['predictions'],
                }
                for p in raw_preds[:10]
            ],
            'day_totals': [],
        }
        if raw_preds:
            dates0 = [pp['date'] for pp in raw_preds[0]['predictions']]
            for i, dt in enumerate(dates0):
                total = sum(p['predictions'][i]['predicted'] for p in raw_preds)
                predict_section['day_totals'].append({'date': dt, 'total': total})
    except Exception as e:
        logger.exception('predict section failed')
        predict_section = None

    # 직전 방역 검증 — 예상(방역 전 7일 평균) 대비 실측(기준일) 편차로 효과 검증 + 3패턴 분류
    # "예상 X → 실측 Y (+Z%)" 형태. 전 관측소 표시(방역 이력 있으면 방역명, 없으면 '방역 이력 없음').
    remedy_verify = []
    try:
        last_plan_by_dev = {}
        for p in all_plans:
            u = p.get('device_uuid'); sd = p.get('scheduled_date') or ''
            if not u or not sd:
                continue
            if u not in last_plan_by_dev or sd > last_plan_by_dev[u]['scheduled_date']:
                m = method_map.get(p.get('method_key'), {})
                last_plan_by_dev[u] = {
                    'scheduled_date': sd, 'method': m.get('name', p.get('method_key')),
                    'worker': p.get('worker') or '', 'volume_l': p.get('volume_l'),
                }
        base_d = sorted_dates[-1] if sorted_dates else None
        # 관측소별 전체 history (backcast용)
        for u, m in name_map.items():
            plan = last_plan_by_dev.get(u)
            seq = [daily[u].get(dt, 0) for dt in sorted_dates[-8:]]
            if len(seq) < 4 or base_d is None:
                continue
            actual = daily[u].get(base_d, 0)                    # 실측 (기준일)
            # 예상 = AI 과거예측(backcast, 방역 미적용). 실패 시 직전 평균으로 폴백.
            full_hist = [{'date': dt, 'count': daily[u].get(dt, 0)} for dt in sorted_dates]
            expected = None
            try:
                bc = predictor.backcast_for_date(full_hist, base_d)
                if bc is not None:
                    expected = float(bc)
            except Exception:
                expected = None
            expected_src = 'AI 예측'
            if expected is None:
                prior = seq[:-1]
                expected = round(sum(prior) / len(prior), 1) if prior else 0
                expected_src = '직전 평균'
            expected = round(expected, 1)
            dev_pct = round((actual - expected) / expected * 100) if expected > 0 else (100 if actual > 0 else 0)
            # 3패턴: 방역 이력 있는 관측소만 판정, 없으면 '검증 대상 외'
            if plan:
                if dev_pct <= 10:
                    pat = {'label': '방역 정상 패턴', 'ver': '효과 양호', 'desc': '방역 후 실측이 예상 이하로 안정 — 정상 효과.', 'action': '현행 방역 주기 유지'}
                elif dev_pct <= 60:
                    pat = {'label': '발생원 누락 의심', 'ver': '효과 부족', 'desc': f'예상 대비 +{dev_pct}% 편차 — 미처리 발생원 잔존 가능성.', 'action': '유충방제(BTi) 추가 + 발생원 재조사'}
                else:
                    pat = {'label': '대형 발생원 미처리 의심', 'ver': '발생원 누락', 'desc': f'예상 대비 +{dev_pct}% 급증 — 대형 발생원 미처리 가능성.', 'action': '발생원 정밀탐색 + 야간 연무(ULV) 긴급 + 유충방제 병행'}
            else:
                pat = {'label': '방역 이력 없음', 'ver': 'N/A', 'desc': '최근 방역 이력이 없어 효과 검증 대상 외.', 'action': '필요 시 방역 계획 등록'}
            remedy_verify.append({
                'name': m['name'],
                'last_method': plan['method'] if plan else '',
                'last_date': plan['scheduled_date'] if plan else '',
                'worker': plan['worker'] if plan else '',
                'volume_l': plan['volume_l'] if plan else None,
                'expected': expected, 'expected_src': expected_src, 'actual': actual, 'change_pct': dev_pct,
                'verdict': pat['ver'], 'pattern': pat['label'], 'desc': pat['desc'], 'action': pat['action'],
                'has_plan': bool(plan),
            })
        # 방역 이력 있는 것 먼저, 편차 큰 순
        remedy_verify.sort(key=lambda x: (x['has_plan'], x['change_pct']), reverse=True)
        # 관측소별 현황 표(stations_section)에 방역이력·효과검증을 병합 — 목업의 통합 표
        verify_by_name = {v['name']: v for v in remedy_verify}
        for s in stations_section:
            v = verify_by_name.get(s['name'])
            if v:
                s['last_method'] = v['last_method']
                s['last_date'] = v['last_date']
                s['worker'] = v['worker']
                s['volume_l'] = v['volume_l']
                s['expected'] = v['expected']
                s['expected_src'] = v['expected_src']
                s['actual'] = v['actual']
                s['change_pct'] = v['change_pct']
                s['verdict'] = v['verdict']
                s['pattern'] = v['pattern']
                s['analysis'] = v['desc']
                s['action'] = v['action']
                s['has_plan'] = v['has_plan']
    except Exception:
        logger.exception('remedy_verify failed')
        remedy_verify = []

    # 일자별 추천 방역 — 예측 등급(모기지수)에 따라 방법·시기·사유 매핑
    remedy_recommend = []
    try:
        if predict_section and predict_section.get('day_totals'):
            # 날짜별 평균 예측지수 (상위 예측 device들의 해당일 predicted_index 평균)
            devs = predict_section.get('devices') or []
            for i, dtot in enumerate(predict_section['day_totals']):
                idx_vals = []
                for d in devs:
                    preds = d.get('predictions') or []
                    if i < len(preds) and preds[i].get('predicted_index') is not None:
                        idx_vals.append(preds[i]['predicted_index'])
                avg_idx = round(sum(idx_vals) / len(idx_vals), 1) if idx_vals else None
                if avg_idx is None:
                    grade = '관심'
                elif avg_idx < 25: grade = '쾌적'
                elif avg_idx < 50: grade = '관심'
                elif avg_idx < 75: grade = '주의'
                else: grade = '불쾌'
                # 등급 → 추천 방역(방법·시기·사유)
                if grade in ('주의', '불쾌'):
                    rec = {'method': 'ULV 야간 연무 + BTi 유충방제', 'timing': '당일 야간 21:00~22:00', 'reason': '성충 급증 예상 — 즉시 살충 + 발생원 차단 병행 필요'}
                elif grade == '관심':
                    rec = {'method': 'BTi 유충방제', 'timing': '오전 09:00~11:00', 'reason': '발생 초기 — 유충 차단이 비용효율적'}
                else:
                    rec = {'method': '예찰 강화(방역 보류)', 'timing': '-', 'reason': '안정 상태 — 추가 방역 불필요'}
                remedy_recommend.append({
                    'date': dtot['date'], 'grade': grade, 'avg_index': avg_idx,
                    'predicted_total': dtot['total'],
                    'method': rec['method'], 'timing': rec['timing'], 'reason': rec['reason'],
                })
    except Exception:
        logger.exception('remedy_recommend failed')
        remedy_recommend = []

    # 권역별 breakdown — 보고서 표용
    region_breakdown = [
        {
            'region_name': rname,
            'total': rd['total'],
            'devices': rd['devices'],
            'top_dev': rd['top_dev'],
            'top_count': rd['top_count'],
            'avg_per_device': round(rd['total'] / rd['devices'], 1) if rd['devices'] else 0,
        }
        for rname, rd in region_sorted
    ]

    # 종합 등급 — 이상감지/경고장비/민원높음 기반 (보고서 메타 표기용)
    _warn = (k.get('warn_count', 0) if k else 0)
    _comp_high = (k.get('complaint_high', 0) if k else 0)
    if len(anomalies) > 0:
        overall_grade = '심각'
    elif _warn > 0:
        overall_grade = '주의'
    elif _comp_high > 0:
        overall_grade = '관심'
    else:
        overall_grade = '안전'

    # ── 섹션별 해석/분석/요약 (파란 박스용) ── 실제 수치 기반 자동 문장
    insights = {}
    try:
        top = (k.get('top_dev') or {}) if k else {}
        chg = k.get('change_pct', 0) if k else 0
        chg_word = '증가' if chg > 5 else ('감소' if chg < -5 else '유지')
        nat = (ov.get('national') if ov else None) or {}
        nat_txt = ''
        if nat.get('avg'):
            nat_txt = f" 전국 일평균 {nat.get('avg')}마리 대비 참고 필요."
        insights['summary'] = (
            f"기준일 전체 포집 {total_in_period}마리(일평균 {avg_per_day}), 전일 대비 {'+' if chg>=0 else ''}{chg}% {chg_word}. "
            f"최다 포집은 {top.get('name') or '-'} {top.get('count', 0)}마리, "
            f"경고 이상 {_warn}개소·이상감지 {len(anomalies)}건.{nat_txt}"
        )
        if anomalies:
            worst = anomalies[0]
            insights['anomaly'] = (
                f"기간 내 {len(anomalies)}건의 이상 감지(≥{ANOMALY_THRESHOLD}마리/일). "
                f"최고는 {worst.get('name') or '-'} {worst.get('count', 0)}마리(임계 {worst.get('bad_min','-')} 대비 +{worst.get('pct_over','-')}%). "
                f"해당 관측소 우선 방역 검토 필요."
            )
        else:
            insights['anomaly'] = f"기간 내 이상 감지(≥{ANOMALY_THRESHOLD}마리/일) 없음 — 전 관측소 안정."
        # 방역 검증 요약
        flagged = [v for v in remedy_verify if v.get('has_plan') and v.get('change_pct', 0) > 10]
        if flagged:
            names = ', '.join(f"{v['name']}(+{v['change_pct']}%)" for v in flagged[:3])
            insights['verify'] = f"방역 후 예상 대비 편차가 큰 {len(flagged)}개소 감지: {names}. 발생원 재조사·추가 방역 권고."
        elif any(v.get('has_plan') for v in remedy_verify):
            insights['verify'] = "방역 실시 관측소 모두 예상 범위 내 — 정상 효과 발휘 중."
        # 예측 요약
        if predict_section and predict_section.get('day_totals'):
            dts = predict_section['day_totals']
            first, last = dts[0]['total'], dts[-1]['total']
            trend_w = '상승' if last > first else ('하강' if last < first else '유지')
            insights['predict'] = f"향후 {len(dts)}일 예측 합계 {first}→{last}마리 {trend_w} 추세. 등급 상향 시 사전 방역 권고."
    except Exception:
        logger.exception('insights build failed')
        insights = {}

    summary = {
        'total_devices': len(devices),
        'insights': insights,
        'period': period, 'period_label': period_label,
        'start_date': start_s, 'end_date': end_s,
        'total_in_period': total_in_period,
        'avg_per_day': avg_per_day,
        'n_days': n_days,
        'overall_grade': overall_grade,
        'anomaly_count': len(anomalies),
        'anomaly_threshold': ANOMALY_THRESHOLD,
        'offline_count': offline_count,
        'low_batt_count': low_batt_count,
        'plan_count': len(plans_in_range),
        'plans_executed_count': plans_executed_count,
        'avg_battery': avg_battery,
        # 기상-생리학
        'avg_temp': avg_temp,
        'avg_humid': avg_humid,
        'avg_precip': avg_precip,
        'temp_zone': temp_zone,
        'humid_zone': humid_zone,
        'trend_text': trend_text,
        'overview_kpi': {
            'today_total': k.get('today_total', 0) if k else 0,
            'change_pct': k.get('change_pct', 0) if k else 0,
            'warn_count': k.get('warn_count', 0) if k else 0,
            'anomaly_today': k.get('anomaly_today', 0) if k else 0,
            'complaint_high': k.get('complaint_high', 0) if k else 0,
            'equip_bad': k.get('equip_bad', 0) if k else 0,
            'equip_lowbatt': k.get('equip_lowbatt', 0) if k else 0,
            'equip_delay': k.get('equip_delay', 0) if k else 0,
            'offline_count': k.get('offline_count', 0) if k else 0,
            'check_count': k.get('check_count', 0) if k else 0,
            'avg_trust': k.get('avg_trust', 0) if k else 0,
            'alert_count': k.get('alert_count', 0) if k else 0,
            'top_dev': k.get('top_dev') if k else None,
        } if k else None,
        'national': (ov.get('national') if ov else None),
        # 풀 섹션
        'sections': {
            'stations': stations_section,
            'heatmap': heatmap_section,
            'anomaly': anomaly_section,
            'predict': predict_section,
            'remedy_verify': remedy_verify,
            'remedy_recommend': remedy_recommend,
            'plans': plans_in_range,
            'region_breakdown': region_breakdown,
            'vector_risks': vector_risks,
            'recent_plans': recent_plans,
            'last7_timeseries': last7_timeseries,
        },
    }
    return report_text, summary, allowed_uuids, source, payload_summary


def moscom_report_api(request):
    """보고서 생성(POST) + 자기 기록 목록(GET)"""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    su = _current_session_user(request)
    if request.method == 'GET':
        # admin은 전체, 일반 사용자는 본인 기록만
        if su.get('is_admin'):
            reports = report_store.list_reports()
        else:
            reports = report_store.list_reports(author_login_id=su.get('login_id'))
        return JsonResponse({'reports': reports})
    try:
        body = _json.loads((request.body or b'').decode('utf-8', errors='replace') or '{}')
    except _json.JSONDecodeError:
        body = {}
    if request.method == 'POST':
        period = body.get('period') or 'daily'
        if period not in ('daily', 'weekly', 'monthly'):
            return JsonResponse({'error': 'period는 daily/weekly/monthly 중 선택'}, status=400)
        base_date = body.get('base_date') or ''
        try:
            report_text, summary, scoped, source, payload = _build_report_body(period, base_date, su, request)
        except Exception as e:
            logger.exception('report body build failed')
            return JsonResponse({'error': f'보고서 데이터 수집 실패: {e}'}, status=500)

        record = {
            'author_login_id': su.get('login_id'),
            'period': period,
            'base_date': summary['end_date'],
            'org': (body.get('org') or '').strip()[:100],
            'department': (body.get('department') or '').strip()[:100],
            'writer_name': (body.get('writer_name') or '').strip()[:50],
            'writer_title': (body.get('writer_title') or '').strip()[:50],
            'reviewer_name': (body.get('reviewer_name') or '').strip()[:50],
            'reviewer_title': (body.get('reviewer_title') or '').strip()[:50],
            'approver_name': (body.get('approver_name') or '').strip()[:50],
            'approver_title': (body.get('approver_title') or '').strip()[:50],
            'summary': summary,
            'report_text': report_text,
            'source': source,
            'scoped_device_uuids': scoped,
        }
        saved = report_store.create_report(record)
        return JsonResponse({'ok': True, 'id': saved['id'], 'report': saved})
    return JsonResponse({'error': 'method not allowed'}, status=405)


def moscom_report_detail_api(request, report_id):
    """단일 보고서 조회 + 삭제(admin 전용)"""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    su = _current_session_user(request)
    rec = report_store.get_report(report_id)
    if not rec:
        return JsonResponse({'error': '존재하지 않는 보고서'}, status=404)
    # 접근 권한: admin 또는 본인이 작성한 것만
    if not su.get('is_admin') and rec.get('author_login_id') != su.get('login_id'):
        return JsonResponse({'error': '권한 없음'}, status=403)
    if request.method == 'GET':
        return JsonResponse({'report': rec})
    if request.method == 'DELETE':
        if not su.get('is_admin'):
            return JsonResponse({'error': 'admin 전용'}, status=403)
        try:
            report_store.delete_report(report_id)
            return JsonResponse({'ok': True})
        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'method not allowed'}, status=405)


def mosquito_report_view(request, report_id):
    """보고서 HTML 페이지 (브라우저에서 열고 인쇄/PDF 저장 가능)"""
    if not request.session.get('mosquito_auth'):
        from django.shortcuts import redirect
        return redirect('/mosquito-test/')
    su = _current_session_user(request)
    rec = report_store.get_report(report_id)
    if not rec:
        return HttpResponse('보고서를 찾을 수 없습니다', status=404)
    if not su.get('is_admin') and rec.get('author_login_id') != su.get('login_id'):
        return HttpResponse('접근 권한이 없습니다', status=403)
    # 히트맵 섹션 → JSON 문자열로 템플릿에 주입
    hm = ((rec.get('summary') or {}).get('sections') or {}).get('heatmap') or []
    return render(request, 'core/mosquito_report.html', {
        'report': rec,
        'heatmap_json': _json.dumps(hm, ensure_ascii=False),
    })


@require_GET
def moscom_admin_judgment(request):
    """AI 행정 판단 리포트 생성 (GPT 연동).
    모든 탭 데이터를 수집해서 요약 → OpenAI에 공문 형식 리포트 요청.
    캐싱 5분.
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    from datetime import datetime, timedelta, timezone
    from django.core.cache import cache
    from django.conf import settings as dj_settings

    su = _current_session_user(request)
    # 사용자별 캐시 키
    cache_key = 'moscom:admin_judgment:' + (su.get('login_id') or 'anon')
    force = request.GET.get('refresh') == '1'
    if not force:
        cached = cache.get(cache_key)
        if cached is not None:
            return JsonResponse(cached)

    try:
        devices = moscom_client.list_devices()
        devices = user_store.filter_devices(su, devices)
        allowed_uuids = {d.get('device_uuid') for d in devices}

        # 7일 통계
        stats = moscom_client.get_statistics(device_uuid='', period_type='2', offset=0)
        stats = [r for r in (stats or []) if r.get('device_uuid') in allowed_uuids]

        # 일별 집계 (장비당)
        from collections import defaultdict
        daily = defaultdict(lambda: defaultdict(int))
        dates = set()
        for r in stats:
            u = r.get('device_uuid')
            date = (r.get('created_date') or '')[:10]
            if u and date:
                daily[u][date] += (r.get('mosquito_count') or 0)
                dates.add(date)
        sorted_dates = sorted(dates)
        # 기준일 = 어제(측정 완료된 날). 오늘은 아직 진행 중이므로 행정 판단 베이스로 부적합.
        try:
            from moscom.timeutil import business_today as _bt
            today_kst_real = _bt().isoformat()
        except Exception:
            today_kst_real = datetime.now(timezone(timedelta(hours=9))).date().isoformat()
        # report_date = 진행 중인 날(today_kst_real) 직전, 즉 어제. sorted_dates 에 어제가 있다면 그것을 사용.
        completed_dates = [d for d in sorted_dates if d < today_kst_real]
        if completed_dates:
            today = completed_dates[-1]   # 측정 완료된 가장 최근일 = 어제
            yday = completed_dates[-2] if len(completed_dates) >= 2 else ''
        else:
            # 어제 데이터가 없으면 폴백: 가장 최근 데이터 사용
            today = sorted_dates[-1] if sorted_dates else ''
            yday = sorted_dates[-2] if len(sorted_dates) >= 2 else ''

        # 기준일 합계 + 그 전일 대비 (오늘은 측정 미완이므로 제외)
        today_total = sum(daily[u].get(today, 0) for u in daily) if today else 0
        yday_total = sum(daily[u].get(yday, 0) for u in daily) if yday else 0
        pct = round((today_total - yday_total) / yday_total * 100) if yday_total else 0

        # 오늘 장비 탑 5
        today_devs = sorted(
            [(u, daily[u].get(today, 0)) for u in daily.keys()],
            key=lambda x: x[1], reverse=True
        )[:5]

        # 장비 이름 매핑 (51마리 전역 임계값)
        ANOMALY_THRESHOLD = 51
        name_map = {}
        for d in devices:
            dv = d.get('device') or {}
            nm = _station_name(dv.get('device_name') or '') or d.get('device_uuid')
            addr = ' '.join(p for p in [dv.get('address_gungu'), dv.get('address_dong')] if p and len(p) < 40) or ''
            name_map[d.get('device_uuid')] = {'name': nm, 'addr': addr, 'bad_min': ANOMALY_THRESHOLD}

        # 51 이상 = '나쁨' / 21~50 = '주의' 인 관측점 (전체 장비 대상)
        bad_stations = []   # 나쁨 (51+)
        warn_stations = []  # 주의 (21~50)
        for u in allowed_uuids:
            v = daily[u].get(today, 0) if today else 0
            if v >= ANOMALY_THRESHOLD:
                bad_stations.append((u, v))
            elif v >= 21:
                warn_stations.append((u, v))
        bad_stations.sort(key=lambda x: -x[1])
        warn_stations.sort(key=lambda x: -x[1])

        # 이상 감지 요약 (51마리 이상)
        anomalies = []
        for u, v in (bad_stations[:10]):
            meta = name_map.get(u, {})
            anomalies.append(f"{meta.get('name')}({meta.get('addr') or '주소미상'}) — 오늘 {v}마리 (기준 {ANOMALY_THRESHOLD} 초과)")

        # 방역 계획 현황
        from core import remedy_store
        visible = allowed_uuids if not su.get('is_admin') else None
        plans = remedy_store.list_plans(visible_uuids=visible)
        method_map = {m['key']: m for m in remedy_store.list_methods()}
        active_plans = []
        for p in plans[:20]:
            m = method_map.get(p['method_key'], {})
            active_plans.append({
                'device': name_map.get(p['device_uuid'], {}).get('name') or p['device_uuid'],
                'method': m.get('name', p['method_key']),
                'scheduled_date': p['scheduled_date'],
                'reduction_pct': m.get('reduction_pct'),
            })

        # 장비 상태 + 온/습도 (전일 대비 증감) — moscom DB 사용
        now_utc = datetime.now(timezone.utc)
        offline_count = 0
        low_batt_count = 0     # < 30% (강준상 요청: 30% 미만)
        for d in devices:
            dv = d.get('device') or {}
            if (dv.get('battery') or 100) < 30:
                low_batt_count += 1
            ud = dv.get('updated_date') or ''
            try:
                udt = datetime.fromisoformat(ud.replace('Z', '+00:00'))
                if (now_utc - udt).total_seconds() / 60 > 1440:
                    offline_count += 1
            except Exception:
                offline_count += 1

        # 온/습도 평균 + 전일 대비 증감 (moscom DB 현재값만 있어서 전일치는 raw 에서 계산)
        avg_temp = avg_humid = None
        temp_delta = humid_delta = None
        try:
            from moscom.models import Device as MoscomDevice, Collection
            from django.utils import timezone as dj_tz
            from datetime import timedelta as td
            md_qs = list(MoscomDevice.objects.filter(device_uuid__in=list(allowed_uuids)))
            temps = [m.temperature for m in md_qs if m.temperature is not None]
            humids = [m.humidity for m in md_qs if m.humidity is not None]
            if temps: avg_temp = round(sum(temps) / len(temps), 1)
            if humids: avg_humid = round(sum(humids) / len(humids), 1)
            # 전일 대비: 24시간 전 raw 의 battery 평균 같은 건 없으나, 온/습도는 실시간 캐시라 어제 자료 없음.
            # 임시: 변화 없음 표시 (None) — Open-Meteo 가 hourly archive 도 지원하지만 별도 호출 필요
        except Exception:
            pass

        # 예측 — moscom_predict 와 같은 로직으로 1주일 예측 합산
        predicted_7d_total = 0
        predicted_7d_avg = 0
        predicted_top = []
        predicted_key_locations = []
        preds_by_uuid = {}
        try:
            from core import predictor
            today_kst_d = datetime.now(timezone(timedelta(hours=9))).date()
            try:
                from moscom.timeutil import business_today
                today_kst_d = business_today()
            except Exception:
                pass
            # 장비별 history
            hist_by_uuid = {u: [] for u in allowed_uuids}
            for r in stats:
                u = r.get('device_uuid')
                date = (r.get('created_date') or '')[:10]
                if u in hist_by_uuid and date and date < today_kst_d.isoformat():
                    hist_by_uuid[u].append({'date': date, 'count': r.get('mosquito_count') or 0})
            for u in hist_by_uuid:
                hist_by_uuid[u].sort(key=lambda h: h['date'])
            inputs = []
            for u in allowed_uuids:
                meta = name_map.get(u, {})
                inputs.append({
                    'uuid': u, 'name': meta.get('name', u), 'region': meta.get('addr', ''),
                    'region_code': '', 'sido': '', 'weather': {},
                    'history': hist_by_uuid[u],
                })
            preds = predictor.predict_for_devices(inputs, days_ahead=7)
            # 일별 합계
            day_totals = defaultdict(int)
            dev_total7 = []
            preds_by_uuid = {}
            for p in preds:
                uuid_ = p.get('uuid')
                total = sum(pp.get('predicted', 0) for pp in p.get('predictions', []))
                max_pred = max((pp.get('predicted', 0) for pp in p.get('predictions', [])), default=0)
                preds_by_uuid[uuid_] = {'predictions': p.get('predictions', []), 'total': total, 'max': max_pred, 'name': p.get('name', '')}
                dev_total7.append((uuid_, p.get('name', ''), total, max_pred))
                for pp in p.get('predictions', []):
                    day_totals[pp.get('date', '')] += pp.get('predicted', 0)
            if day_totals:
                predicted_7d_total = sum(day_totals.values())
                predicted_7d_avg = round(predicted_7d_total / len(day_totals))
            # 주요 위치 = 7일 누적 상위 5 + 어제 실측 상위 3 의 합집합
            top_by_pred = sorted(dev_total7, key=lambda x: -x[2])[:5]
            top_uuids_set = {x[0] for x in top_by_pred}
            for u, _v in today_devs[:3]:
                if u not in top_uuids_set:
                    pb = preds_by_uuid.get(u)
                    if pb:
                        top_by_pred.append((u, pb['name'], pb['total'], pb['max']))
                        top_uuids_set.add(u)
            predicted_top = top_by_pred
        except Exception as e:
            logger.warning(f'predict in admin_judgment failed: {e}')
            preds_by_uuid = {}

        # ── 추가 분석: 권역별 / 7일 트렌드 / 매개체 위험 / 방역 이력 효과 ──
        try:
            from moscom.models import Device as MoscomDevice, Region as MoscomRegion, Collection
            region_name_by_code = {r.code: r.name for r in MoscomRegion.objects.all()}
            uuid_to_md = {md.device_uuid: md for md in MoscomDevice.objects.all()}
        except Exception:
            region_name_by_code, uuid_to_md = {}, {}

        # 권역별 오늘 합계 + 평균
        region_today = {}  # region_name -> {total, count, devices:[{name, count}]}
        for u in allowed_uuids:
            md = uuid_to_md.get(u)
            rc = (md.region_code if md else '') or ''
            rname = region_name_by_code.get(rc, rc) or '미지정'
            v = daily[u].get(today, 0) if today else 0
            if rname not in region_today:
                region_today[rname] = {'total': 0, 'count': 0, 'devices': []}
            region_today[rname]['total'] += v
            region_today[rname]['count'] += 1
            region_today[rname]['devices'].append({
                'name': name_map.get(u, {}).get('name', u),
                'count': v,
            })
        region_today_sorted = sorted(
            ((k, v) for k, v in region_today.items()),
            key=lambda kv: -kv[1]['total']
        )

        # 7일 시계열 (전 장비 합계) — 측정 완료된 날만 포함
        last7_daily = []
        ref_dates = completed_dates if completed_dates else sorted_dates
        for d in ref_dates[-7:]:
            day_sum = sum(daily[u].get(d, 0) for u in allowed_uuids)
            last7_daily.append({'date': d, 'total': day_sum})
        # 추세 텍스트
        trend_arrow = ''
        if len(last7_daily) >= 4:
            half = len(last7_daily) // 2
            first_half_avg = sum(x['total'] for x in last7_daily[:half]) / max(1, half)
            second_half_avg = sum(x['total'] for x in last7_daily[half:]) / max(1, len(last7_daily) - half)
            if first_half_avg > 0:
                pct_trend = round((second_half_avg - first_half_avg) / first_half_avg * 100)
                if pct_trend > 15: trend_arrow = f'7일 추세: 전반 일평균 {round(first_half_avg)} → 후반 {round(second_half_avg)} (+{pct_trend}%, 상승)'
                elif pct_trend < -15: trend_arrow = f'7일 추세: 전반 일평균 {round(first_half_avg)} → 후반 {round(second_half_avg)} ({pct_trend}%, 하강)'
                else: trend_arrow = f'7일 추세: 전반 {round(first_half_avg)} ↔ 후반 {round(second_half_avg)} (안정)'

        # 매개체 위험 평가 (기상 기반)
        vector_risks = []
        try:
            from moscom.health_knowledge import assess_vector_risks, thermal_zone_for, humidity_zone_for
            recent_rainfall = sum((m.precipitation or 0) for m in uuid_to_md.values() if m.device_uuid in allowed_uuids) / max(1, len([m for m in uuid_to_md.values() if m.device_uuid in allowed_uuids]))
            vector_risks = assess_vector_risks(avg_temp, avg_humid, recent_rainfall)
            temp_zone = thermal_zone_for(avg_temp)
            humid_zone = humidity_zone_for(avg_humid)
        except Exception as e:
            logger.warning(f'vector risk assess failed: {e}')
            temp_zone = humid_zone = None

        # 방역 효과 (최근 14일 내 실시 건수, 누적 감소 추정)
        from datetime import timedelta as _td
        recent_plans = []
        plan_count_recent = 0
        try:
            today_kst = datetime.now(timezone(timedelta(hours=9))).date()
            from datetime import date as _date_cls
            for p in plans:
                sd = p.get('scheduled_date') or ''
                try:
                    pd_ = _date_cls.fromisoformat(sd)
                    days_ago = (today_kst - pd_).days
                    if 0 <= days_ago <= 14:
                        plan_count_recent += 1
                        m = method_map.get(p.get('method_key'), {})
                        recent_plans.append({
                            'device': name_map.get(p['device_uuid'], {}).get('name') or p['device_uuid'],
                            'method': m.get('name', p.get('method_key', '')),
                            'date': sd,
                            'reduction_pct': m.get('reduction_pct'),
                            'days_ago': days_ago,
                        })
                except ValueError:
                    continue
        except Exception:
            pass

        # ── GPT 프롬프트 구성 ────────────────────────
        top_devs_text = '\n'.join(
            f"  - {name_map.get(u, {}).get('name')} ({name_map.get(u, {}).get('addr') or '주소미상'}): {v}마리"
            for u, v in today_devs
        ) or '  - 데이터 없음'
        anomaly_text = '\n'.join(f"  - {a}" for a in anomalies) or '  - 51마리 초과 장비 없음'
        bad_text = '\n'.join(
            f"  - {name_map.get(u, {}).get('name')} ({name_map.get(u, {}).get('addr') or '주소미상'}): {v}마리 — 나쁨"
            for u, v in bad_stations[:10]
        ) or '  - 없음'
        warn_text = '\n'.join(
            f"  - {name_map.get(u, {}).get('name')} ({name_map.get(u, {}).get('addr') or '주소미상'}): {v}마리 — 주의"
            for u, v in warn_stations[:10]
        ) or '  - 없음'
        plans_text = '\n'.join(
            f"  - {p['device']} / {p['method']} / 실시 {p['scheduled_date']} / 감소율 {p['reduction_pct']}%"
            for p in active_plans
        ) or '  - 등록된 방역 실시 내역 없음'
        # 주요 위치 예측 상세 — 권역/주소/일별 예측/등급 포함 (전체 합 대신 핵심)
        def _grade_pred(n):
            if n <= 10: return '안전'
            if n <= 50: return '관심'
            if n <= 100: return '주의'
            if n <= 200: return '경고'
            return '위험'
        pred_top_lines = []
        predicted_key_locations = []   # 화면/요약용
        for uuid_, nm, tot, mx in predicted_top[:8]:
            # 권역명
            md = uuid_to_md.get(uuid_)
            rc = (md.region_code if md else '') or ''
            rname = region_name_by_code.get(rc, rc) or '미지정'
            addr = name_map.get(uuid_, {}).get('addr', '') or '주소미상'
            pb = preds_by_uuid.get(uuid_, {})
            daily_preds = pb.get('predictions', [])
            yday_actual = daily[uuid_].get(today, 0) if today else 0
            # 7일 일별 표시 (최대 7개)
            day_str = ' / '.join(f"{pp.get('date','')[-5:]}:{pp.get('predicted',0)}" for pp in daily_preds[:7])
            grade = _grade_pred(mx)
            pred_top_lines.append(
                f"  - {nm} [{rname}] ({addr}) — 어제 실측 {yday_actual}마리 → 7일 누적 예측 {tot}마리 (최대 {mx}, 등급 {grade}) | {day_str}"
            )
            predicted_key_locations.append({
                'name': nm, 'region': rname, 'addr': addr,
                'yday_actual': yday_actual,
                'total_7d': tot, 'max_7d': mx, 'grade': grade,
                'predictions': [
                    {'date': pp.get('date', ''), 'predicted': pp.get('predicted', 0)}
                    for pp in daily_preds[:7]
                ],
            })
        pred_top_text = '\n'.join(pred_top_lines) or '  - 예측 데이터 없음'

        # 예측 vs 전일 변화 — '안정세' 오판 방지
        if yday_total > 0 and predicted_7d_avg > 0:
            pred_vs_yday_ratio = round(predicted_7d_avg / yday_total, 1)
        else:
            pred_vs_yday_ratio = 0
        # 자동 판정 가이드 (GPT에 전달)
        if pred_vs_yday_ratio >= 1.5:
            pred_trend_hint = f'⚠ 향후 7일 일평균({predicted_7d_avg}마리)이 전일({yday_total}마리) 대비 {pred_vs_yday_ratio:.1f}배 — 급증 예상'
        elif pred_vs_yday_ratio >= 1.1:
            pred_trend_hint = f'△ 향후 7일 일평균({predicted_7d_avg}마리)이 전일({yday_total}마리) 대비 {pred_vs_yday_ratio:.1f}배 — 증가 예상'
        elif pred_vs_yday_ratio <= 0.7 and pred_vs_yday_ratio > 0:
            pred_trend_hint = f'▽ 향후 7일 일평균({predicted_7d_avg}마리)이 전일({yday_total}마리) 대비 {pred_vs_yday_ratio:.1f}배 — 감소 예상'
        elif predicted_7d_avg > 0:
            pred_trend_hint = f'≈ 향후 7일 일평균({predicted_7d_avg}마리)이 전일({yday_total}마리) 대비 {pred_vs_yday_ratio:.1f}배 — 안정세'
        else:
            pred_trend_hint = '예측 데이터 부족'

        date_label = today or datetime.now(timezone(timedelta(hours=9))).strftime('%Y-%m-%d')
        temp_disp = f'{avg_temp}°C' if avg_temp is not None else '데이터 없음'
        humid_disp = f'{avg_humid}%' if avg_humid is not None else '데이터 없음'

        # 권역별 텍스트
        region_text = '\n'.join(
            f"  - {rname}: {info['total']}마리 (장비 {info['count']}대, 평균 {round(info['total']/max(1,info['count']),1)}마리)"
            for rname, info in region_today_sorted[:10]
        ) or '  - 권역 데이터 없음'

        # 7일 시계열
        last7_text = ' / '.join(f"{d['date'][-5:]}: {d['total']}" for d in last7_daily) or '데이터 없음'

        # 매개체 위험
        vector_text = '\n'.join(
            f"  - {v['species']}: 위험 {v['risk_score']}점 ({v['risk_level']}) "
            f"매개질병={','.join(v['diseases'])} 활동시간={v['active_hours']} 서식지={v['breeding_site']}"
            for v in vector_risks[:4]
        ) or '  - 매개체 평가 데이터 없음'

        # 기상-우화 영역
        thermal_text = ''
        if temp_zone:
            thermal_text = f"  - 기온 영역: {temp_zone['range'][0]}~{temp_zone['range'][1]}°C → {temp_zone['effect']} (유충→성충 {temp_zone['larval_days']}일, 성충 수명 {temp_zone['adult_lifespan']})"
        humid_text = ''
        if humid_zone:
            humid_text = f"  - 습도 영역: {humid_zone['range'][0]}~{humid_zone['range'][1]}% → {humid_zone['effect']}"

        # 최근 방역 이력
        recent_plans_text = '\n'.join(
            f"  - {p['date']} ({p['days_ago']}일 전): {p['device']} / {p['method']} (감소율 {p['reduction_pct']}%)"
            for p in recent_plans[:8]
        ) or '  - 최근 14일 내 시행된 방역 없음'

        payload_summary = f"""[모기 발생 감시 종합 보고 · 기준일 {date_label} (어제, 측정 완료)]
※ 오늘은 아직 측정 진행 중이므로 행정 판단은 측정 완료된 어제({date_label}) 데이터를 기준으로 함.

■ 전체 수치 요약
  - 전체 운영 장비: {len(devices)}대
  - 기준일({date_label}) 포집 합계: {today_total}마리 (그 전일 {yday_total}마리, {pct:+d}%)
  - 평균 기온: {temp_disp}
  - 평균 습도: {humid_disp}
  - 오프라인 장비: {offline_count}대
  - 배터리 30% 미만 장비: {low_batt_count}대 (현장 점검 대상)

■ 권역별 포집 현황 (기준일 · Top 10)
{region_text}

■ 기준일 상위 포집 장비 (Top 5)
{top_devs_text}

■ 방역 상태 '나쁨' 관측점 (51마리↑, 즉시 방역 필요)
{bad_text}

■ 방역 상태 '주의' 관측점 (21~50마리, 감시 강화)
{warn_text}

■ 7일 시계열 (측정 완료된 날만 · 전 장비 합계)
  {last7_text}
  {trend_arrow}

■ 기상-모기 생리학 분석
{thermal_text}
{humid_text}

■ 매개체 위험 평가 (현 기상 조건 기반)
{vector_text}

■ AI 예측 — 주요 위치 중심 분석
※ 전체 합계보다 '핵심 관측점별' 향후 7일 궤적이 행정 판단에 더 유용함.
  - 핵심 관측점 정의: 7일 누적 예측 상위 5 + 기준일 실측 상위 3 (합집합)
  - (참고) 전체 합 일평균 {predicted_7d_avg}마리 / 누적 {predicted_7d_total}마리 / 전일대비 {pred_vs_yday_ratio:.1f}배 ({pred_trend_hint})

■ 핵심 관측점별 7일 예측 (위치·등급·일별 궤적)
{pred_top_text}

■ 최근 14일 시행된 방역 이력
{recent_plans_text}

■ 등록된 방역 계획
{plans_text}
"""

        # OpenAI 호출
        api_key = getattr(dj_settings, 'OPENAI_API_KEY', '') or ''
        if not api_key:
            report_text = '※ OpenAI API 키가 서버에 설정되지 않아 AI 분석이 불가합니다.\n\n' + payload_summary
            source = 'fallback'
        else:
            try:
                import openai
                client = openai.OpenAI(api_key=api_key)
                resp = client.chat.completions.create(
                    model='gpt-4o-mini',
                    messages=[
                        {'role': 'system', 'content': (
                            '당신은 지자체 보건소 방역 담당자에게 "왜 이런 방역을 해야 하는지" 그 근거를 설명하는 '
                            '모기 매개체·방역 전문가 AI입니다. 한국어로 작성하십시오.\n'
                            '\n'
                            '핵심 목표: 일반적인 현황 나열이 아니라, "이 관측소에는 이 방역을 이 시점에 해야 한다"는 '
                            '추천과, 그 추천에 이르게 된 **인과적 근거(데이터 → 원인 → 추천)**를 명확히 설명하는 것입니다. '
                            '담당자가 보고서를 읽고 "아, 그래서 이 방역을 추천하는구나"라고 납득할 수 있어야 합니다.\n'
                            '\n'
                            '== 기준일 원칙 ==\n'
                            '"기준일"은 측정이 완료된 전일(어제)입니다. 시점은 "전일/금일"로 표기하십시오.\n'
                            '\n'
                            '== 보고서 구조 (다음 4개 섹션, 순서대로) ==\n'
                            '■ 1) 현재 상황 요약\n'
                            '   - 기준일 전체/권역 포집 추세, 위험 관측소(나쁨·주의)를 마릿수·전일 대비 변화와 함께 2~4문장으로 요약.\n'
                            '\n'
                            '■ 2) 방역이 필요한 이유 (근거)\n'
                            '   - **이 보고서의 핵심.** 왜 지금 방역이 필요한지를 데이터→원인의 인과로 설명.\n'
                            '   - 포집 근거: "○○공원 75마리(전일 +40%), 인접 권역 평균 22마리의 3배" 처럼 비교 수치로.\n'
                            '   - 기상-생리 근거: 현 기온/습도가 우화·성충 생존·산란에 미치는 영향을 연결.\n'
                            '     (25~28°C→우화 5~7일 / 습도 75%↑→성충 수명 연장 / 강수 후 5~7일 1차 우화 피크)\n'
                            '   - 매개체 근거: 데이터의 매개체 위험 점수를 인용 ("작은빨간집모기 78점(높음)—일본뇌염 매개").\n'
                            '   - AI 예측 근거: 핵심 관측소의 향후 7일 추이(상승/유지)·최대 예측일·등급으로 "방치 시 어떻게 되는지" 제시.\n'
                            '\n'
                            '■ 3) 관측소별 추천 방역\n'
                            '   - 위험·주의 관측소마다: "[관측소명] → [추천 방역] : [이 방역을 고른 이유]" 형식으로.\n'
                            '   - 방역별 선택 근거(반드시 명시):\n'
                            '     Bti 살포 = 유충 표적, 48시간 내 90% 사망 (수변부·유충 발생지)\n'
                            '     ULV 초미립자 연무 = 성충 즉시 살충, 18~22시 야간 시행 효과 최대 (성충 급증지)\n'
                            '     잔류분무 = 14~21일 지속 효과 (재발 우려지)\n'
                            '     용기제거 = 흰줄숲모기 95% 감소 (주택가·용기 다수지)\n'
                            '   - 즉, "왜 이 관측소에 ULV인가 / 왜 Bti인가"가 드러나야 함.\n'
                            '\n'
                            '■ 4) 우선순위 권고\n'
                            '   - 어디부터 먼저 방역할지 우선순위 불릿 2~4개. 가장 시급한 관측소·방역·시점.\n'
                            '\n'
                            '== 규칙 ==\n'
                            '- 모든 추천에 근거 수치를 붙이고, "안정세"는 전일 대비 ±10% 이내일 때만 사용(1.5배↑는 "급증").\n'
                            '- 각 섹션은 "■ N) 섹션명"으로 시작, 불릿은 "  - "로 시작.\n'
                            '- 어조는 간결한 행정체(…함, …필요함, …권고함). 불필요한 미사여구·중복 금지.\n'
                            '- 결재란·서명란·날짜 줄은 포함하지 마세요 (별도 렌더링).'
                        )},
                        {'role': 'user', 'content': payload_summary},
                    ],
                    temperature=0.4,
                    max_tokens=1800,
                )
                report_text = resp.choices[0].message.content.strip()
                source = 'openai:gpt-4o-mini'
            except Exception as e:
                logger.exception('OpenAI call failed')
                report_text = f'※ AI 분석 중 오류가 발생했습니다: {e}\n\n' + payload_summary
                source = 'error'

        result = {
            'report_text': report_text,
            'source': source,
            'generated_at': datetime.now(timezone(timedelta(hours=9))).isoformat(),
            'author': {'login_id': su.get('login_id'), 'is_admin': su.get('is_admin')},
            'summary': {
                'total_devices': len(devices),
                'today_total': today_total,
                'yday_total': yday_total,
                'change_pct': pct,
                'anomaly_count': len(anomalies),
                'bad_count': len(bad_stations),     # 51마리 이상 (나쁨)
                'warn_count': len(warn_stations),   # 21~50 (주의)
                'offline_count': offline_count,
                'low_batt_count': low_batt_count,   # 30% 미만
                'plan_count': len(active_plans),
                'avg_temperature': avg_temp,
                'avg_humidity': avg_humid,
                'predicted_7d_total': predicted_7d_total,
                'predicted_7d_avg': predicted_7d_avg,
                # 신규 분석 데이터 (화면 표시용)
                'base_date': today,           # 기준일 (어제, 측정 완료)
                'base_date_label': '기준일(전일, 측정 완료)',
                'region_breakdown': [
                    {'region_name': rn, 'total': info['total'], 'device_count': info['count'],
                     'avg': round(info['total']/max(1,info['count']),1)}
                    for rn, info in region_today_sorted[:10]
                ],
                'last7_timeseries': last7_daily,
                'vector_risks': vector_risks,
                'recent_plans': recent_plans,
                'predicted_key_locations': predicted_key_locations,
            },
        }
        cache.set(cache_key, result, 600)
        return JsonResponse(result)
    except Exception as e:
        logger.exception('admin_judgment failed')
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def moscom_equipment_health(request):
    """장비 신뢰도 점수화 (4축)
    - 배터리: 현재 battery%
    - 팬 작동: fan (1=정상, 0=정지)
    - 수신 지연: 현재 시각 - updated_date (분)
    - 포집 0 지속: 최근 3일간 mosquito_count 모두 0
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    from datetime import datetime, timedelta, timezone
    from collections import defaultdict
    try:
        now = datetime.now(timezone.utc)
        devices = moscom_client.list_devices()
        devices = user_store.filter_devices(_current_session_user(request), devices)

        # 최근 3일 일별 통계
        stats = moscom_client.get_statistics(device_uuid='', period_type='2', offset=0)
        daily = defaultdict(lambda: defaultdict(int))
        for r in (stats or []):
            u = r.get('device_uuid')
            date = (r.get('created_date') or '')[:10]
            if u and date:
                daily[u][date] += (r.get('mosquito_count') or 0)

        # 최근 1시간 수신 횟수 — raw 통계 레코드 수로 집계 (관측소별)
        recv_1h = defaultdict(int)
        try:
            h1_start = (now - timedelta(hours=1)).strftime('%Y-%m-%dT%H:%M:%S.000Z')
            h1_end = now.strftime('%Y-%m-%dT%H:%M:%S.000Z')
            raw1h = moscom_client.get_statistics_by_date(
                start_dt=h1_start, end_dt=h1_end, aggregation='raw', device_uuid='0',
            ) or []
            for r in raw1h:
                u = r.get('device_uuid')
                if u:
                    recv_1h[u] += 1
        except Exception as e:
            logger.warning(f'equipment recv_1h failed: {e}')

        def parse_iso(s):
            if not s:
                return None
            try:
                s2 = s.replace('Z', '+00:00')
                return datetime.fromisoformat(s2)
            except Exception:
                return None

        results = []
        for d in devices:
            u = d.get('device_uuid')
            dv = d.get('device') or {}
            name = _station_name(dv.get('device_name') or '') or u
            battery = dv.get('battery') or 0
            fan = dv.get('fan') or 0
            charge = dv.get('charge') or 0
            updated = parse_iso(dv.get('updated_date'))

            # 축 1: 배터리 (0~100점)
            if battery >= 50: axis_battery = 100
            elif battery >= 30: axis_battery = 80
            elif battery >= 20: axis_battery = 60
            elif battery >= 10: axis_battery = 35
            else: axis_battery = 10
            # 축 2: 팬
            axis_fan = 100 if fan == 1 else 30

            # 축 3: 수신 지연
            if updated:
                delay_min = (now - updated).total_seconds() / 60
            else:
                delay_min = 99999
            if delay_min <= 120: axis_signal = 100     # 2h 이내
            elif delay_min <= 360: axis_signal = 80    # 6h
            elif delay_min <= 720: axis_signal = 55    # 12h
            elif delay_min <= 1440: axis_signal = 30   # 24h
            else: axis_signal = 5                      # 24h 초과

            # 축 4: 포집 0 지속 — 최근 3일치 합
            dev_daily = daily.get(u, {})
            recent_vals = list(dev_daily.values())
            total3 = sum(recent_vals[-3:]) if recent_vals else 0
            if len(recent_vals) >= 3 and total3 == 0:
                axis_zero = 40  # 3일 내내 0 → 센서 의심
            elif total3 == 0:
                axis_zero = 65  # 데이터 부족
            else:
                axis_zero = 100

            # 가중치: 배터리 25 / 팬 20 / 수신 35 / 포집0 20
            trust_score = round(
                axis_battery * 0.25 + axis_fan * 0.20 + axis_signal * 0.35 + axis_zero * 0.20, 1
            )

            # 등급
            if trust_score >= 85: trust = 'HIGH'
            elif trust_score >= 60: trust = 'MEDIUM'
            else: trust = 'LOW'

            # 장비 상태 판정
            if axis_signal <= 30:
                status = '오프라인'
            elif trust_score < 60 or battery < 15 or (fan == 0 and axis_signal >= 55):
                status = '점검필요'
            else:
                status = '정상'

            # 이상 원인 수집
            issues = []
            if battery < 20: issues.append(f'배터리 부족 ({battery}%)')
            if fan == 0: issues.append('팬 정지')
            if delay_min > 720: issues.append(f'수신 지연 {int(delay_min/60)}시간')
            elif delay_min > 360: issues.append(f'수신 지연 {int(delay_min/60)}시간')
            if axis_zero == 40: issues.append('3일 연속 포집 0 (센서 의심)')

            # 조치 안내
            if status == '오프라인':
                action = '현장 방문 · 전원·통신 점검'
            elif battery < 15:
                action = '배터리 교체 필요'
            elif fan == 0:
                action = '팬 점검 · 재기동'
            elif axis_zero == 40:
                action = '센서부 청소 점검'
            elif trust == 'HIGH':
                action = '정상 운영 중'
            else:
                action = '원격 상태 모니터링'

            parts = [dv.get('address_gungu'), dv.get('address_dong')]
            addr = ' '.join(p for p in parts if p and len(p) < 40 and any('\uac00' <= c <= '\ud7a3' or c.isalnum() or c == ' ' for c in p)).strip()

            results.append({
                'uuid': u,
                'name': name,
                'address': addr,
                'battery': battery,
                'fan': fan,
                'charge': charge,
                'updated_date': dv.get('updated_date'),
                'delay_minutes': round(delay_min, 1) if delay_min < 99999 else None,
                'recv_1h': recv_1h.get(u, 0),
                'status': status,
                'trust': trust,
                'trust_score': trust_score,
                'axes': {
                    '배터리': axis_battery,
                    '팬': axis_fan,
                    '수신': axis_signal,
                    '포집': axis_zero,
                },
                'issues': issues,
                'action': action,
            })

        results.sort(key=lambda r: (r['trust_score'], r['delay_minutes'] or 0))

        # 집계
        summary = {
            'total': len(results),
            'normal': sum(1 for r in results if r['status'] == '정상'),
            'check': sum(1 for r in results if r['status'] == '점검필요'),
            'offline': sum(1 for r in results if r['status'] == '오프라인'),
            'avg_trust': round(sum(r['trust_score'] for r in results) / len(results), 1) if results else 0,
        }

        return JsonResponse({
            'count': len(results),
            'summary': summary,
            'criteria': {
                'weights': {'배터리': 25, '팬': 20, '수신': 35, '포집': 20},
                'trust': {'HIGH': '≥ 85', 'MEDIUM': '60~84', 'LOW': '< 60'},
            },
            'items': results,
        })
    except Exception as e:
        logger.exception('equipment_health failed')
        return JsonResponse({'error': str(e)}, status=500)


def _visible_uuids_for(request):
    """세션 사용자의 허용 장비 UUID 집합. admin=None(=전부)."""
    su = _current_session_user(request)
    if not su:
        return set()
    if su.get('is_admin'):
        return None
    return set(su.get('allowed_devices') or [])


@require_GET
def moscom_remedy_methods(request):
    """방역 방법 목록 반환"""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    return JsonResponse({'methods': remedy_store.list_methods()})


def moscom_remedy_api(request):
    """방역 계획 목록/추가
    GET  : 세션 사용자의 허용 장비 한정 목록
    POST : 새 계획 추가 (body: device_uuid, method_key, scheduled_date, note)
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    visible = _visible_uuids_for(request)
    if request.method == 'GET':
        plans = remedy_store.list_plans(visible_uuids=visible)
        return JsonResponse({'plans': plans, 'methods': remedy_store.list_methods()})
    try:
        body = _json.loads((request.body or b'').decode('utf-8', errors='replace') or '{}')
    except _json.JSONDecodeError:
        body = {}
    if request.method == 'POST':
        dev_uuid = body.get('device_uuid')
        if visible is not None and dev_uuid not in visible:
            return JsonResponse({'error': '허용되지 않은 장비입니다'}, status=403)
        try:
            su = _current_session_user(request)
            plan = remedy_store.create_plan(
                owner_id=su.get('login_id', ''),
                device_uuid=dev_uuid,
                method_key=body.get('method_key'),
                method_keys=body.get('method_keys'),
                scheduled_date=body.get('scheduled_date'),
                note=body.get('note', ''),
                worker=body.get('worker', ''),
                volume_l=body.get('volume_l'),
            )
            return JsonResponse({'ok': True, 'plan': plan})
        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'method not allowed'}, status=405)


def moscom_remedy_detail_api(request, plan_id):
    """방역 계획 수정/삭제"""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    plan = remedy_store.get_plan(plan_id)
    if not plan:
        return JsonResponse({'error': '해당 계획을 찾을 수 없습니다'}, status=404)
    visible = _visible_uuids_for(request)
    if visible is not None and plan.get('device_uuid') not in visible:
        return JsonResponse({'error': '권한 없음'}, status=403)
    try:
        body = _json.loads((request.body or b'').decode('utf-8', errors='replace') or '{}')
    except _json.JSONDecodeError:
        body = {}
    try:
        if request.method in ('PUT', 'PATCH'):
            # device_uuid 변경 시도 시 권한 재확인
            new_uuid = body.get('device_uuid')
            if new_uuid and visible is not None and new_uuid not in visible:
                return JsonResponse({'error': '허용되지 않은 장비입니다'}, status=403)
            p = remedy_store.update_plan(plan_id, body)
            return JsonResponse({'ok': True, 'plan': p})
        if request.method == 'DELETE':
            remedy_store.delete_plan(plan_id)
            return JsonResponse({'ok': True})
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'method not allowed'}, status=405)


@require_GET
def moscom_remedy_template(request):
    """방역 실시 내역 일괄등록용 엑셀 템플릿(.xlsx) 다운로드."""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = '방역실시내역'
    headers = ['대상 관측소', '방역 방법1(필수)', '방역 방법2', '방역 방법3', '실시일(YYYY-MM-DD)', '비고']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1B3A6B')
    # 예시 행
    method_names = [m['name'] for m in remedy_store.list_methods()]
    ws.append(['관측소명 예: 정자공원', method_names[0] if method_names else '', '없음', '없음', '2026-06-14', '예시 — 삭제 후 입력'])
    # 안내 시트(유효 방역명 목록)
    ws2 = wb.create_sheet('방역방법 목록')
    ws2.append(['사용 가능한 방역 방법명 (그대로 복사해 입력)'])
    for n in method_names:
        ws2.append([n])
    ws2.append(['없음 (방법2·3 비울 때)'])
    ws.column_dimensions['A'].width = 22
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 24
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 24
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="remedy_template.xlsx"'
    return resp


@csrf_exempt
@require_POST
def moscom_remedy_import(request):
    """엑셀(.xlsx) 업로드 → 방역 실시 내역 일괄 등록.
    컬럼: 대상 관측소 / 방역 방법1(필수) / 방역 방법2 / 방역 방법3 / 실시일 / 비고
    관측소명·한글 방역명으로 매칭. 행별 결과 반환.
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': '파일이 없습니다'}, status=400)
    su = _current_session_user(request)
    visible = _visible_uuids_for(request)

    # 관측소명 → uuid 맵 (정제명/원본명 모두 키로)
    devices = moscom_client.list_devices()
    devices = user_store.filter_devices(su, devices)
    name_to_uuid = {}
    for d in devices:
        dv = d.get('device') or {}
        u = d.get('device_uuid')
        for nm in {(dv.get('device_name') or '').strip(), _station_name(dv.get('device_name') or '')}:
            if nm:
                name_to_uuid[nm] = u
    # 한글 방역명 → key 맵
    name_to_method = {m['name']: m['key'] for m in remedy_store.list_methods()}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({'error': f'엑셀 읽기 실패: {e}'}, status=400)

    def norm_method(v):
        s = (v or '').strip()
        if not s or s in ('없음', '-', 'N/A', 'na'):
            return None
        return name_to_method.get(s, '__INVALID__' if s else None)

    results = {'ok': 0, 'fail': 0, 'errors': []}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows, start=2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        cells = list(row) + [None] * (6 - len(row))
        station = (str(cells[0]).strip() if cells[0] is not None else '')
        m1, m2, m3 = norm_method(cells[1]), norm_method(cells[2]), norm_method(cells[3])
        sched = cells[4]
        note = (str(cells[5]).strip() if cells[5] is not None else '')
        # 예시 행 스킵
        if station.startswith('관측소명 예'):
            continue
        # 실시일 정규화
        if hasattr(sched, 'strftime'):
            sched = sched.strftime('%Y-%m-%d')
        else:
            sched = (str(sched).strip()[:10] if sched is not None else '')
        # 검증
        uuid_ = name_to_uuid.get(station)
        if not uuid_:
            results['fail'] += 1; results['errors'].append(f'{i}행: 관측소 "{station}" 매칭 실패'); continue
        if visible is not None and uuid_ not in visible:
            results['fail'] += 1; results['errors'].append(f'{i}행: 권한 없는 관측소 "{station}"'); continue
        if m1 is None or m1 == '__INVALID__':
            results['fail'] += 1; results['errors'].append(f'{i}행: 방역 방법1이 비었거나 잘못됨'); continue
        mks = [m1] + [m for m in (m2, m3) if m and m != '__INVALID__']
        try:
            remedy_store.create_plan(
                owner_id=su.get('login_id', ''), device_uuid=uuid_,
                method_key=mks[0], method_keys=mks, scheduled_date=sched, note=note,
            )
            results['ok'] += 1
        except ValueError as e:
            results['fail'] += 1; results['errors'].append(f'{i}행: {e}')
    return JsonResponse(results)


@require_GET
def kakao_status(request):
    """현재 로그인 사용자의 카카오 연동 상태"""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    su = _current_session_user(request)
    entry = kakao_client.get_token(su.get('login_id'))
    if not entry:
        return JsonResponse({'connected': False})
    return JsonResponse({
        'connected': True,
        'connected_at': entry.get('connected_at'),
        'access_expires_at': entry.get('access_expires_at'),
        'scopes': entry.get('scopes') or [],
    })


def kakao_oauth_start(request):
    """카카오 OAuth 시작 — 로그인 페이지로 리다이렉트"""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    from django.shortcuts import redirect
    url = kakao_client.authorize_url()
    return redirect(url)


def kakao_oauth_callback(request):
    """카카오가 code와 함께 돌아오는 지점"""
    if not request.session.get('mosquito_auth'):
        from django.shortcuts import redirect
        return redirect('/mosquito-test/')
    su = _current_session_user(request)
    code = request.GET.get('code')
    err = request.GET.get('error')
    if err:
        return HttpResponse(f'<h3>카카오 인증 실패</h3><p>{err}: {request.GET.get("error_description","")}</p><p><a href="/mosquito-test/">돌아가기</a></p>', status=400)
    if not code:
        return HttpResponse('code 없음', status=400)
    try:
        payload = kakao_client.exchange_code(code)
        kakao_client.save_tokens_for(su.get('login_id'), payload)
    except Exception as e:
        logger.exception('kakao exchange failed')
        return HttpResponse(f'토큰 교환 실패: {e}', status=500)
    # 완료 후 관리자 탭으로 돌려보냄
    return HttpResponse(
        '<!DOCTYPE html><html><head><meta charset="utf-8"><title>연동 완료</title></head>'
        '<body style="font-family:sans-serif;text-align:center;padding:50px">'
        '<h2 style="color:#1B3A6B">카카오톡 연동 완료</h2>'
        '<p>이제 대시보드에서 "나에게 알림 보내기"를 사용할 수 있습니다.</p>'
        '<p><a href="/mosquito-test/" style="color:#2980b9">대시보드로 돌아가기</a></p>'
        '<script>if(window.opener){window.opener.postMessage({type:"kakao-connected"},"*");setTimeout(()=>window.close(),1500)}</script>'
        '</body></html>'
    )


def kakao_disconnect(request):
    """연동 해제 (토큰 삭제)"""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    su = _current_session_user(request)
    kakao_client.delete_token(su.get('login_id'))
    return JsonResponse({'ok': True})


def kakao_send_api(request):
    """나에게 보내기"""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    try:
        body = _json.loads((request.body or b'').decode('utf-8', errors='replace') or '{}')
    except _json.JSONDecodeError:
        body = {}
    text = (body.get('text') or '').strip()
    link = body.get('link') or 'https://saerong.com/mosquito-test/'
    if not text:
        return JsonResponse({'error': '메시지 내용이 비어있습니다'}, status=400)
    su = _current_session_user(request)
    ok, detail = kakao_client.send_to_me(su.get('login_id'), text, link_url=link)
    if not ok:
        return JsonResponse({'error': str(detail)}, status=400)
    return JsonResponse({'ok': True, 'result': detail})


def moscom_user_detail_api(request, login_id):
    admin_err = _require_admin(request)
    if admin_err:
        return admin_err
    try:
        body = _json.loads((request.body or b'').decode('utf-8', errors='replace') or '{}')
    except _json.JSONDecodeError:
        body = {}
    try:
        if request.method in ('PUT', 'PATCH'):
            user_store.update_user(
                login_id=login_id,
                password=body.get('password') if body.get('password') else None,
                allowed_devices=body.get('allowed_devices') if 'allowed_devices' in body else None,
            )
            return JsonResponse({'ok': True, 'users': user_store.list_users()})
        if request.method == 'DELETE':
            user_store.delete_user(login_id)
            return JsonResponse({'ok': True, 'users': user_store.list_users()})
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'method not allowed'}, status=405)


@require_GET
def moscom_devices(request):
    """MOSCOM API 장비 목록 프록시 (허용 장비만 반환).
    응답의 각 장비에 region_code / region_name 주입 (moscom.Device + moscom.Region 조회).
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    try:
        force = request.GET.get('refresh') == '1'
        data = moscom_client.list_devices(force_refresh=force)
        # 권역 매핑
        try:
            from moscom.models import Device as MoscomDevice, Region as MoscomRegion
            region_name_by_code = {r.code: r.name for r in MoscomRegion.objects.all()}
            md_by_uuid = {md.device_uuid: md for md in MoscomDevice.objects.only('device_uuid', 'region_code')}
        except Exception:
            region_name_by_code, md_by_uuid = {}, {}

        def _inject(entry):
            u = entry.get('device_uuid') or (entry.get('device') or {}).get('device_uuid')
            md = md_by_uuid.get(u)
            rc = (md.region_code if md else '') or ''
            entry['region_code'] = rc
            entry['region_name'] = region_name_by_code.get(rc, rc) or '미지정'
            return entry

        data = [_inject(e) for e in (data or [])]

        # 관리자 세션일 때 '전체' 플래그로 원본 반환 (관리자 탭 사용자 추가에서 전 장비 리스트 필요)
        if request.GET.get('all') == '1' and bool(request.session.get('mosquito_is_admin')):
            return JsonResponse({'count': len(data), 'devices': data, 'admin_all': True}, safe=False)
        su = _current_session_user(request)
        filtered = user_store.filter_devices(su, data)
        return JsonResponse({'count': len(filtered), 'devices': filtered}, safe=False)
    except Exception as e:
        logger.exception('MOSCOM /device/listAll failed')
        return JsonResponse({'error': str(e)}, status=502)


@require_GET
def moscom_raw_collection(request):
    """MOSCOM API 기간별 포집 데이터 프록시
    쿼리스트링: start, end, device_uuid(선택)
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    start = request.GET.get('start')
    end = request.GET.get('end')
    if not start or not end:
        return JsonResponse({'error': 'start, end query params required (ISO 8601)'}, status=400)
    try:
        data = moscom_client.raw_collection_bulk(
            start_dt=start,
            end_dt=end,
            device_uuid=request.GET.get('device_uuid') or None,
        )
        return JsonResponse({'data': data}, safe=False)
    except Exception as e:
        logger.exception('MOSCOM /device/rawCollectionBulk failed')
        return JsonResponse({'error': str(e)}, status=502)


def _filter_records_by_uuid(request, records):
    """세션 사용자의 허용 장비로 레코드 배열 필터. admin은 그대로."""
    allowed = user_store.allowed_uuid_set(_current_session_user(request))
    if allowed is None:
        return records
    return [r for r in records if r.get('device_uuid') in allowed]


@require_GET
def moscom_statistics(request):
    """MOSCOM API 장비별 일별 통계 프록시
    쿼리스트링: period(0=3개월, 1=4주, 2=7일, 3=기타, 기본=2), device_uuid(선택, 빈값=전체)
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    try:
        period = request.GET.get('period', '2')
        device_uuid = request.GET.get('device_uuid', '')
        offset = int(request.GET.get('offset', '0'))
        data = moscom_client.get_statistics(
            device_uuid=device_uuid, period_type=period, offset=offset,
        )
        data = _filter_records_by_uuid(request, data or [])
        return JsonResponse({'count': len(data), 'stats': data}, safe=False)
    except Exception as e:
        logger.exception('MOSCOM /device/statistics failed')
        return JsonResponse({'error': str(e)}, status=502)


@require_GET
def moscom_daily(request):
    """MOSCOM API 일별 집계 프록시 — 7일 추세 탭용 (임의 기간)
    쿼리스트링: start, end (YYYY-MM-DD, KST). 최대 90일.
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    from datetime import datetime, timedelta, timezone
    try:
        start_str = request.GET.get('start')
        end_str = request.GET.get('end')
        if not start_str or not end_str:
            return JsonResponse({'error': 'start, end query params required (YYYY-MM-DD)'}, status=400)
        try:
            start_d = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_d = datetime.strptime(end_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'invalid date format, use YYYY-MM-DD'}, status=400)
        if end_d < start_d:
            start_d, end_d = end_d, start_d
        if (end_d - start_d).days > 90:
            return JsonResponse({'error': 'range too large (max 90 days)'}, status=400)

        # MOSCOM의 "업무일"은 KST 10:00 ~ 다음날 09:59 수집.
        # 선택한 날짜 D는 "D일 10:00 KST ~ (D+1)일 10:00 KST" 로 해석.
        # start_utc = start_d 10:00 KST (-> UTC 01:00)
        # end_utc   = (end_d + 1) 10:00 KST (-> 다음날 UTC 01:00) exclusive
        start_utc = datetime(start_d.year, start_d.month, start_d.day, 10, 0, 0, tzinfo=timezone.utc) - timedelta(hours=9)
        end_utc = datetime(end_d.year, end_d.month, end_d.day, 10, 0, 0, tzinfo=timezone.utc) + timedelta(days=1) - timedelta(hours=9)
        start_iso = start_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')
        end_iso = end_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')

        data = moscom_client.get_statistics_by_date(
            start_dt=start_iso, end_dt=end_iso, aggregation='day', device_uuid='0',
        )
        data = _filter_records_by_uuid(request, data or [])
        return JsonResponse({
            'count': len(data) if isinstance(data, list) else 0,
            'start': start_str, 'end': end_str,
            'data': data,
        }, safe=False)
    except Exception as e:
        logger.exception('MOSCOM /device/statisticsByDate (day) failed')
        return JsonResponse({'error': str(e)}, status=502)


def _valid_kor(s):
    """한글/영문/숫자/공백만 포함한 정상 문자열인지 (깨진 주소 필터링)"""
    if not s or len(s) > 60:
        return False
    return any('\uac00' <= c <= '\ud7a3' or c.isalnum() or c == ' ' for c in s) and not any('\ud800' <= c <= '\udfff' or c == '\ufffd' for c in s)


@require_GET
@require_GET
def moscom_anomaly_history(request):
    """이상 감지 누적 이력. 최근 N일(기본 30일) 중 bad_min 초과한 일자×장비만 반환.
    쿼리스트링: days (기본 30, 최대 90)
    응답: {items: [{date, uuid, name, addr, count, bad_min, pct_over}]}
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    from datetime import datetime, timedelta, timezone
    from collections import defaultdict
    try:
        try:
            days = int(request.GET.get('days', '30'))
        except (TypeError, ValueError):
            days = 30
        days = max(1, min(days, 90))

        su = _current_session_user(request)
        devices = moscom_client.list_devices()
        devices = user_store.filter_devices(su, devices)
        meta = {}
        for d in devices:
            u = d.get('device_uuid')
            dv = d.get('device') or {}
            name = _station_name(dv.get('device_name') or '') or u
            addr = ' '.join(p for p in [dv.get('address_gungu'), dv.get('address_dong')] if p and len(p) < 40 and _valid_kor(p)).strip()
            bad_min = 50  # 경보 기준 50마리 고정 (이상감지 탭과 통일)
            meta[u] = {'name': name, 'addr': addr, 'bad_min': bad_min}

        # 최근 N일 일별 통계
        now = datetime.now(timezone.utc)
        end_kst = now + timedelta(hours=9)
        start_d = (end_kst - timedelta(days=days - 1)).date()
        start_utc = datetime(start_d.year, start_d.month, start_d.day, 10, 0, 0, tzinfo=timezone.utc) - timedelta(hours=9)
        end_utc = now
        stats = moscom_client.get_statistics_by_date(
            start_dt=start_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z'),
            end_dt=end_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z'),
            aggregation='day', device_uuid='0',
        )
        # 일별 합산
        daily = defaultdict(lambda: defaultdict(int))
        for r in (stats or []):
            u = r.get('device_uuid')
            date = (r.get('created_date') or '')[:10]
            if u in meta and date:
                daily[u][date] += (r.get('mosquito_count') or 0)

        items = []
        for u, m in meta.items():
            bm = m.get('bad_min', 100) or 100
            for date, cnt in daily.get(u, {}).items():
                if cnt >= bm and bm > 0:
                    items.append({
                        'date': date,
                        'uuid': u,
                        'name': m['name'],
                        'addr': m['addr'],
                        'count': cnt,
                        'bad_min': bm,
                        'pct_over': round((cnt - bm) / bm * 100),
                    })
        # 최신순
        items.sort(key=lambda x: (x['date'], x['count']), reverse=True)
        return JsonResponse({
            'days': days,
            'count': len(items),
            'items': items,
        })
    except Exception as e:
        logger.exception('anomaly_history failed')
        return JsonResponse({'error': str(e)}, status=500)


def moscom_complaint_risk(request):
    """민원 가능 지역 위험 점수 산출 (4축 가중합)
    쿼리스트링: date=YYYY-MM-DD (선택, 미지정시 최신일)
    위험 점수: 절대 포집량 40% + 증가율 25% + 추세 20% + 전국 대비 15%
    민원 점수: 체감 포집량 35% + 급증 신호 25% + 야간 피크 25% + 주거지 인접 15%
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    from datetime import datetime, timedelta, timezone
    from collections import defaultdict
    try:
        # 0) 쿼리 파라미터: 기준일 (선택)
        target_date = (request.GET.get('date') or '').strip()
        if target_date:
            try:
                datetime.strptime(target_date, '%Y-%m-%d')
            except ValueError:
                return JsonResponse({'error': 'date 형식은 YYYY-MM-DD'}, status=400)

        # 1) 장비 메타
        devices = moscom_client.list_devices()
        devices = user_store.filter_devices(_current_session_user(request), devices)
        # 수동 지정 2축 분류 lookup (uuid → (region_type, form_type))
        habitat_map = {}
        try:
            from moscom.models import Device as MoscomDevice
            for md in MoscomDevice.objects.all():
                habitat_map[md.device_uuid] = (md.region_type or '', md.form_type or '')
        except Exception:
            pass
        meta = {}
        for d in devices:
            u = d.get('device_uuid')
            dv = d.get('device') or {}
            name = _station_name(dv.get('device_name') or '') or u
            sido = (dv.get('address_sido') or '').strip()
            gungu = (dv.get('address_gungu') or '').strip()
            dong = (dv.get('address_dong') or '').strip()
            detail = (dv.get('address_detail') or '').strip()
            bad_min = ((dv.get('deviceSetting') or {}).get('bad_min')) or 100
            rt, ft = habitat_map.get(u, ('', ''))
            meta[u] = {
                'name': name,
                'sido': sido if _valid_kor(sido) else '',
                'gungu': gungu if _valid_kor(gungu) else '',
                'dong': dong if _valid_kor(dong) else '',
                'detail': detail if _valid_kor(detail) else '',
                'bad_min': bad_min,
                'region_type': rt,
                'form_type': ft,
            }

        # 2) 7일 통계 (장비별 일별 포집량)
        if target_date:
            # 임의 기준일 → 해당일 포함 7일 (statisticsByDate)
            target_d = datetime.strptime(target_date, '%Y-%m-%d').date()
            start_d = target_d - timedelta(days=6)
            # 업무일(KST 10시) 커버: start 10:00 KST ~ (target+1) 10:00 KST
            s_utc = datetime(start_d.year, start_d.month, start_d.day, 10, 0, 0, tzinfo=timezone.utc) - timedelta(hours=9)
            e_utc = datetime(target_d.year, target_d.month, target_d.day, 10, 0, 0, tzinfo=timezone.utc) + timedelta(days=1) - timedelta(hours=9)
            stats = moscom_client.get_statistics_by_date(
                start_dt=s_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z'),
                end_dt=e_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z'),
                aggregation='day', device_uuid='0',
            )
        else:
            stats = moscom_client.get_statistics(device_uuid='', period_type='2', offset=0)

        daily = defaultdict(lambda: defaultdict(int))
        dates_set = set()
        for r in (stats or []):
            u = r.get('device_uuid')
            date = (r.get('created_date') or '')[:10]
            if not u or not date or u not in meta:
                continue
            # target_date가 지정된 경우 그 이후 날짜는 제외 (UI 일관성)
            if target_date and date > target_date:
                continue
            daily[u][date] += (r.get('mosquito_count') or 0)
            dates_set.add(date)
        all_dates = sorted(dates_set)
        if target_date:
            today = target_date
            # yday는 today 직전 날
            prev_dates = [d for d in all_dates if d < today]
            yday = prev_dates[-1] if prev_dates else ''
        else:
            # 종합현황과 동일: 측정 완료된 전일(업무일 어제) 기준. 오늘은 수집 진행 중이라 제외
            try:
                from moscom.timeutil import business_yesterday
                by = business_yesterday().isoformat()
            except Exception:
                by = ''
            completed = [d for d in all_dates if not by or d <= by]
            today = (completed[-1] if completed else (all_dates[-1] if all_dates else ''))
            prev_dates = [d for d in all_dates if d < today]
            yday = prev_dates[-1] if prev_dates else ''
        week_dates = [d for d in all_dates if d <= today][-7:]

        # 3) 최근 48h raw 데이터 → 장비별 야간 피크 비율
        now = datetime.now(timezone.utc)
        start_iso = (now - timedelta(days=2)).strftime('%Y-%m-%dT%H:%M:%S.000Z')
        end_iso = now.strftime('%Y-%m-%dT%H:%M:%S.000Z')
        raw = moscom_client.get_statistics_by_date(start_dt=start_iso, end_dt=end_iso, aggregation='raw', device_uuid='0')
        per_dev_hour = defaultdict(lambda: [None] * 24)
        for r in (raw or []):
            u = r.get('device_uuid')
            if u not in meta:
                continue
            iso = r.get('created_date') or ''
            try:
                utc_h = int(iso[11:13])
            except Exception:
                continue
            kst_h = (utc_h + 9) % 24
            cnt = r.get('mosquito_count') or 0
            cur = per_dev_hour[u][kst_h]
            if cur is None or cnt > cur:
                per_dev_hour[u][kst_h] = cnt
        dev_night_ratio = {}
        for u, hr in per_dev_hour.items():
            deltas = [0] * 24
            prev = 0
            for h in range(24):
                v = hr[h]
                if v is None:
                    continue
                d = v - prev
                deltas[h] = d if d > 0 else 0
                prev = v
            night = sum(deltas[19:24])
            total = sum(deltas)
            dev_night_ratio[u] = (night / total) if total > 0 else 0.0

        # 4) 오늘 백분위
        today_counts = sorted([daily[u].get(today, 0) for u in meta.keys()], reverse=True)
        n = len(today_counts)
        def percentile_rank(v):
            if n <= 1:
                return 100
            above = sum(1 for x in today_counts if x < v)
            return round((above / (n - 1)) * 100)

        # 5) 주거지 키워드
        residential_keywords = ['공원', '아파트', '주거', '학교', '초등', '중학', '고등', '어린이집', '유치원']
        def residential_hit(m):
            txt = ' '.join([m['sido'], m['gungu'], m['dong'], m['detail'], m['name']])
            return any(k in txt for k in residential_keywords)

        # 6) 장비별 점수 계산
        results = []
        for u, m in meta.items():
            today_cnt = daily[u].get(today, 0) if today else 0
            yday_cnt = daily[u].get(yday, 0) if yday else 0
            week_vals = [daily[u].get(d, 0) for d in week_dates]
            week_avg = sum(week_vals) / len(week_vals) if week_vals else 0

            axis1 = min(100, (today_cnt / m['bad_min']) * 100) if m['bad_min'] > 0 else 0
            if yday_cnt > 0:
                axis2 = max(0, min(100, (today_cnt - yday_cnt) / yday_cnt * 100))
            else:
                axis2 = 50 if today_cnt > 0 else 0
            if len(week_vals) >= 4:
                half = len(week_vals) // 2
                fh = sum(week_vals[:half]) / half if half > 0 else 0
                lh = sum(week_vals[half:]) / (len(week_vals) - half)
                if fh > 0:
                    tr = (lh - fh) / fh * 100
                    axis3 = max(0, min(100, (tr / 30) * 100))
                else:
                    axis3 = 50 if lh > 0 else 0
            else:
                axis3 = 0
            axis4 = percentile_rank(today_cnt)

            risk_score = round(axis1 * 0.40 + axis2 * 0.25 + axis3 * 0.20 + axis4 * 0.15, 1)
            if risk_score <= 20: risk_level = '안전'
            elif risk_score <= 40: risk_level = '관심'
            elif risk_score <= 60: risk_level = '주의'
            elif risk_score <= 80: risk_level = '경고'
            else: risk_level = '심각'

            night_ratio = dev_night_ratio.get(u, 0.0)
            axis_felt = axis1
            axis_surge = axis2
            axis_night = min(100, (night_ratio / 0.6) * 100) if night_ratio else 0
            resi_hit = residential_hit(m)
            axis_resi = 80 if resi_hit else 20

            complaint_score = round(
                axis_felt * 0.35 + axis_surge * 0.25 + axis_night * 0.25 + axis_resi * 0.15, 1
            )
            if complaint_score <= 30: complaint_level = '낮음'
            elif complaint_score <= 60: complaint_level = '보통'
            else: complaint_level = '높음'

            # 주요 원인 (민원 기여도 상위 2개)
            axes_complaint = [
                ('체감 포집량', axis_felt, f'기준치(bad_min {m["bad_min"]}) 대비 {round(today_cnt / max(m["bad_min"],1) * 100)}%'),
                ('급증 신호', axis_surge, f'어제 {yday_cnt}마리 → 오늘 {today_cnt}마리'),
                ('야간 피크', axis_night, f'19~23시 비중 {round(night_ratio*100)}%'),
                ('주거지 인접', axis_resi, '주거지 키워드 매칭' if resi_hit else '주거지 키워드 없음'),
            ]
            axes_complaint.sort(key=lambda x: x[1], reverse=True)
            causes = [{'name': a[0], 'score': round(a[1], 1), 'detail': a[2]} for a in axes_complaint[:2]]

            actions = []
            if complaint_score >= 61:
                actions.append('긴급 유충 방제 + 주민 사전 안내문 배포')
            if risk_score >= 61 and axis3 >= 60:
                actions.append('성충 방제 + 24시간 모니터링')
            if axis_night >= 60:
                actions.append('19~23시 집중 방제')
            if not actions:
                actions.append('정기 예찰 유지')

            region = ' '.join(p for p in [m['sido'], m['gungu']] if p) or '기타'

            results.append({
                'uuid': u,
                'name': m['name'],
                'region': region,
                'address': ' '.join(p for p in [m['gungu'], m['dong'], m['detail']] if p),
                'region_type': m.get('region_type') or '',
                'form_type': m.get('form_type') or '',
                'bad_min': m['bad_min'],
                'today_count': today_cnt,
                'yday_count': yday_cnt,
                'week_avg': round(week_avg, 1),
                'night_ratio': round(night_ratio, 3),
                'residential': resi_hit,
                'risk': {
                    'score': risk_score, 'level': risk_level,
                    'axes': {
                        '절대 포집량': round(axis1, 1),
                        '증가율': round(axis2, 1),
                        '추세': round(axis3, 1),
                        '전국 대비': round(axis4, 1),
                    },
                },
                'complaint': {
                    'score': complaint_score, 'level': complaint_level,
                    'axes': {
                        '체감 포집량': round(axis_felt, 1),
                        '급증 신호': round(axis_surge, 1),
                        '야간 피크': round(axis_night, 1),
                        '주거지 인접': round(axis_resi, 1),
                    },
                },
                'causes': causes,
                'actions': actions,
            })

        results.sort(key=lambda r: (r['complaint']['score'], r['risk']['score']), reverse=True)

        return JsonResponse({
            'count': len(results),
            'today': today,
            'criteria': {
                'risk': {
                    'weights': {'절대 포집량': 40, '증가율': 25, '추세': 20, '전국 대비': 15},
                    'levels': {'안전': '0~20', '관심': '21~40', '주의': '41~60', '경고': '61~80', '심각': '81~100'},
                },
                'complaint': {
                    'weights': {'체감 포집량': 35, '급증 신호': 25, '야간 피크': 25, '주거지 인접': 15},
                    'levels': {'낮음': '0~30', '보통': '31~60', '높음': '61~100'},
                    'residential_keywords': residential_keywords,
                },
            },
            'items': results,
        }, safe=False)
    except Exception as e:
        logger.exception('complaint_risk failed')
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def moscom_predict(request):
    """AI 모기 발생 예측 (내일~3일 후, 장비 52개 전체)
    학습 데이터: 청주 3개소·여름 2개월. 다른 지역/계절은 참고값.
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    from datetime import datetime, timedelta, timezone
    try:
        # 최근 10일 일별 통계 (lag7까지 쓰니 넉넉하게)
        # 영업일 기준 오늘 (새벽 5시 경계)
        try:
            from moscom.timeutil import business_today
            today_kst = business_today()
        except Exception:
            today_kst = datetime.now(timezone(timedelta(hours=9))).date()
        start_d = today_kst - timedelta(days=10)
        start_iso = datetime(start_d.year, start_d.month, start_d.day, 0, 0, 0, tzinfo=timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')
        end_iso = (datetime(today_kst.year, today_kst.month, today_kst.day, 0, 0, 0, tzinfo=timezone.utc) + timedelta(days=1)).strftime('%Y-%m-%dT%H:%M:%S.000Z')

        daily = moscom_client.get_statistics_by_date(start_dt=start_iso, end_dt=end_iso, aggregation='day', device_uuid='0')
        devices = moscom_client.list_devices()
        # 세션 사용자 허용 장비로 제한
        devices = user_store.filter_devices(_current_session_user(request), devices)

        # 장비 메타 + moscom DB (region_code, 날씨)
        moscom_device_map = {}
        region_name_by_code = {}
        try:
            from moscom.models import Device as MoscomDevice, Region as MoscomRegion
            region_name_by_code = {r.code: r.name for r in MoscomRegion.objects.all()}
            for md in MoscomDevice.objects.all():
                moscom_device_map[md.device_uuid] = md
        except Exception:
            pass

        meta = {}
        for d in devices:
            u = d.get('device_uuid')
            dv = d.get('device') or {}
            name = _station_name(dv.get('device_name') or '') or u
            md = moscom_device_map.get(u)
            rcode = (md.region_code if md else '') or ''
            # region 그룹 키: 권역명 (KH→김해 본시 등) 우선, 없으면 시도+군구
            if rcode and region_name_by_code.get(rcode):
                region = region_name_by_code[rcode]
            else:
                parts = [dv.get('address_sido'), dv.get('address_gungu')]
                region = ' '.join(p for p in parts if p and len(p) < 40 and any(ord(c) < 0x3400 or 0xAC00 <= ord(c) <= 0xD7A3 for c in p)) or '기타'
            meta[u] = {
                'name': name, 'region': region,
                'region_code': rcode,
                'sido': (md.address_sido if md else (dv.get('address_sido') or '')),
                'weather': {
                    'temperature': md.temperature if md else None,
                    'humidity': md.humidity if md else None,
                    'precipitation': md.precipitation if md else None,
                    'wind_speed': md.wind_speed if md else None,
                } if md else {},
            }

        # 장비별 history 생성
        # 오늘(기준일)은 아직 측정이 안 끝났으니 실측에 포함하지 않음 (예측 컬럼과 중복 방지)
        today_iso = today_kst.isoformat()
        hist_by_uuid = {u: [] for u in meta}
        for r in (daily or []):
            u = r.get('device_uuid')
            if u not in hist_by_uuid:
                continue
            date = (r.get('created_date') or '')[:10]
            if not date:
                continue
            # 오늘 이후 데이터는 실측에서 제외
            if date >= today_iso:
                continue
            hist_by_uuid[u].append({'date': date, 'count': r.get('mosquito_count') or 0})

        # 히스토리는 오름차순(과거 → 최신)으로 정렬: 차트와 lag 계산에 필요
        for u in hist_by_uuid:
            hist_by_uuid[u].sort(key=lambda h: h['date'])

        # 예측 대상 입력
        inputs = []
        for u, m in meta.items():
            inputs.append({
                'uuid': u,
                'name': m['name'],
                'region': m['region'],
                'history': hist_by_uuid[u],
                'region_code': m.get('region_code') or '',
                'sido': m.get('sido') or '',
                'weather': m.get('weather') or {},
            })

        try:
            days_ahead = int(request.GET.get('days', '3'))
        except (TypeError, ValueError):
            days_ahead = 3
        preds = predictor.predict_for_devices(inputs, days_ahead=days_ahead)

        # 방역 계획 효과 반영 (post-processing) — predicted_index/grade 보존
        def _grade_count(n):
            if n <= 10: return '안전'
            if n <= 50: return '관심'
            if n <= 100: return '주의'
            if n <= 200: return '경고'
            return '위험'
        def _grade_idx(v):
            if v is None: return None
            if v < 25: return '쾌적'
            if v < 50: return '관심'
            if v < 75: return '주의'
            return '불쾌'
        remedy_summary_by_uuid = {}
        for p in preds:
            uid = p['uuid']
            new_preds = []
            applied_by_date = {}
            for pp in p['predictions']:
                factor, applied = remedy_store.adjustment_factor(uid, pp.get('date'))
                orig = pp.get('predicted') or 0
                adj = int(round(orig * factor))
                # 모기지수도 방역 계수만큼 비례 감소 (마릿수가 줄면 지수도 줄어듦)
                orig_idx = pp.get('predicted_index')
                if orig_idx is not None:
                    adj_idx = round(max(0.0, min(100.0, orig_idx * factor)), 1)
                else:
                    adj_idx = None
                new_preds.append({
                    'date': pp['date'],
                    'predicted': adj,
                    'predicted_raw': orig,
                    'predicted_index': adj_idx,
                    'predicted_index_raw': round(orig_idx, 1) if orig_idx is not None else None,
                    'grade': _grade_idx(adj_idx),
                    'remedy_factor': round(factor, 3),
                })
                if applied:
                    applied_by_date[pp['date']] = applied
            p['predictions'] = new_preds
            ps = [x['predicted'] for x in new_preds]
            idxs = [x['predicted_index'] for x in new_preds if x['predicted_index'] is not None]
            p['max_predicted'] = max(ps) if ps else 0
            p['avg_predicted'] = round(sum(ps) / len(ps)) if ps else 0
            if idxs:
                avg_idx = sum(idxs) / len(idxs)
                p['max_index'] = round(max(idxs), 1)
                p['avg_index'] = round(avg_idx, 1)
                p['grade'] = _grade_idx(avg_idx)
            else:
                p['max_index'] = None
                p['avg_index'] = None
                p['grade'] = _grade_count(p['max_predicted'])
            if applied_by_date:
                p['remedy_applied'] = applied_by_date
                remedy_summary_by_uuid[uid] = applied_by_date
                # 추론 근거에 방역 반영 요약 덧붙임
                _names = []
                for _alist in applied_by_date.values():
                    for _a in (_alist or []):
                        _nm = (_a.get('method_name') or _a.get('name')) if isinstance(_a, dict) else str(_a)
                        if _nm and _nm not in _names:
                            _names.append(_nm)
                if _names:
                    _msg = f"방역 {len(applied_by_date)}일 반영(예: {_names[0]})"
                    p['reasoning'] = (p.get('reasoning') or '') + ' · ' + _msg if p.get('reasoning') else _msg

        # max_predicted 내림차순 정렬
        preds.sort(key=lambda x: x.get('max_predicted', 0), reverse=True)
        return JsonResponse({
            'count': len(preds),
            'model': 'RandomForest',
            'remedy_applied_count': len(remedy_summary_by_uuid),
            'model_info': {
                'name': '예측 모델 v1.0',
                'algorithm': 'Random Forest Regressor',
                'n_trees': 200, 'max_depth': 10,
                'trained_on': '3개 관측소 · 2025년 7~8월 · 187건',
                'features': 34,
            },
            'disclaimer': '학습 데이터: 3개 관측소 · 2025년 7~8월. 다른 지역/계절은 참고값.',
            'predictions': preds,
        }, safe=False)
    except Exception as e:
        logger.exception('AI predict failed')
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def moscom_forecast_brief(request):
    """AI 위험도 예보 — 보건소/질병청용 종합 브리핑.
    1) 권역별 위험 신호등 (3일 평균 모기지수 기준)
    2) 기상 → 모기 발생 과학적 설명 (GPT 자동 생성 + 룰 fallback)
    3) 트렌드 비교 (최근 7일 vs 직전 7일 vs 작년 동기간)
    4) 방역 시뮬레이션은 별도 endpoint (moscom_forecast_simulate)
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    from datetime import datetime, timedelta, timezone, date as date_cls
    from collections import defaultdict
    from django.core.cache import cache
    from django.conf import settings as dj_settings

    su = _current_session_user(request)
    cache_key = 'moscom:forecast_brief:' + (su.get('login_id') or 'anon')
    force = request.GET.get('refresh') == '1'
    if not force:
        cached = cache.get(cache_key)
        if cached is not None:
            return JsonResponse(cached)

    try:
        # 영업일 today
        try:
            from moscom.timeutil import business_today, business_yesterday
            today_d = business_today()
            yday_d = business_yesterday()
        except Exception:
            today_d = datetime.now(timezone(timedelta(hours=9))).date()
            yday_d = today_d - timedelta(days=1)

        devices = moscom_client.list_devices()
        devices = user_store.filter_devices(su, devices)
        allowed_uuids = {d.get('device_uuid') for d in devices}

        # moscom DB lookup (region + weather)
        moscom_device_map = {}
        region_name_by_code = {}
        try:
            from moscom.models import Device as MoscomDevice, Region as MoscomRegion, Collection
            region_name_by_code = {r.code: r.name for r in MoscomRegion.objects.all()}
            for md in MoscomDevice.objects.all():
                moscom_device_map[md.device_uuid] = md
        except Exception:
            Collection = None  # noqa

        # 장비별 history 7일
        stats = moscom_client.get_statistics(device_uuid='', period_type='2', offset=0)
        stats = [r for r in (stats or []) if r.get('device_uuid') in allowed_uuids]
        daily = defaultdict(lambda: defaultdict(int))
        for r in stats:
            u = r.get('device_uuid')
            date = (r.get('created_date') or '')[:10]
            if u and date and date < today_d.isoformat():
                daily[u][date] += (r.get('mosquito_count') or 0)

        # 예측 호출 (3일)
        try:
            from core import predictor
            inputs = []
            for u in allowed_uuids:
                hist = [{'date': d, 'count': c} for d, c in sorted(daily[u].items())]
                md = moscom_device_map.get(u)
                name = _station_name(md.device_name if md else '') or u
                inputs.append({
                    'uuid': u, 'name': name,
                    'region': ' '.join(p for p in [(md.address_sido if md else ''), (md.address_gungu if md else '')] if p),
                    'region_code': (md.region_code if md else '') or '',
                    'sido': (md.address_sido if md else '') or '',
                    'history': hist,
                    'weather': {
                        'temperature': md.temperature if md else None,
                        'humidity': md.humidity if md else None,
                        'precipitation': md.precipitation if md else None,
                        'wind_speed': md.wind_speed if md else None,
                    } if md else {},
                })
            preds = predictor.predict_for_devices(inputs, days_ahead=3)
        except Exception as e:
            logger.warning(f'forecast_brief predict failed: {e}')
            preds = []

        # ── 1. 권역별 위험 신호등 ──
        region_groups = defaultdict(list)  # region_name -> [{uuid, name, avg_index, max_predicted}]
        for p in preds:
            md = moscom_device_map.get(p.get('uuid'))
            rc = (md.region_code if md else '') or ''
            rname = region_name_by_code.get(rc, rc) or '미지정'
            region_groups[rname].append({
                'uuid': p.get('uuid'),
                'name': p.get('name'),
                'avg_index': p.get('avg_index'),
                'max_index': p.get('max_index'),
                'max_predicted': p.get('max_predicted'),
                'grade': p.get('grade'),
            })
        signals = []
        for rname, items in region_groups.items():
            vals = [it.get('avg_index') for it in items if it.get('avg_index') is not None]
            if not vals:
                continue
            avg_idx = sum(vals) / len(vals)
            max_dev = max(items, key=lambda x: (x.get('avg_index') or 0))
            grade = '쾌적' if avg_idx < 25 else '관심' if avg_idx < 50 else '주의' if avg_idx < 75 else '불쾌'
            color = {'쾌적': 'green', '관심': 'yellow', '주의': 'orange', '불쾌': 'red'}[grade]
            signals.append({
                'region_name': rname,
                'device_count': len(items),
                'avg_index': round(avg_idx, 1),
                'grade': grade,
                'color': color,
                'top_device': {
                    'name': max_dev.get('name'),
                    'avg_index': max_dev.get('avg_index'),
                    'max_predicted': max_dev.get('max_predicted'),
                },
            })
        signals.sort(key=lambda x: -x['avg_index'])

        # ── 1-2. 관측소별 위험 신호등 (9p: 권역이 아닌 관측소 단위) ──
        stations = []
        _grade_color = {'쾌적': 'green', '관심': 'yellow', '주의': 'orange', '불쾌': 'red'}
        for p in preds:
            md = moscom_device_map.get(p.get('uuid'))
            rc = (md.region_code if md else '') or ''
            rname = region_name_by_code.get(rc, rc) or '미지정'
            ai = p.get('avg_index')
            grade = p.get('grade') or (
                '쾌적' if (ai or 0) < 25 else '관심' if ai < 50 else '주의' if ai < 75 else '불쾌'
            )
            stations.append({
                'uuid': p.get('uuid'),
                'name': p.get('name'),
                'region_name': rname,
                'avg_index': round(ai, 1) if ai is not None else None,
                'max_index': p.get('max_index'),
                'max_predicted': p.get('max_predicted'),
                'grade': grade,
                'color': _grade_color.get(grade, 'green'),
                'reasoning': p.get('reasoning') or '',
            })
        stations.sort(key=lambda x: -(x.get('avg_index') or 0))

        # ── 2. 기상 평균 (전체 장비 기준) ──
        temps = [m.temperature for m in moscom_device_map.values() if m.device_uuid in allowed_uuids and m.temperature is not None]
        humids = [m.humidity for m in moscom_device_map.values() if m.device_uuid in allowed_uuids and m.humidity is not None]
        precs = [m.precipitation for m in moscom_device_map.values() if m.device_uuid in allowed_uuids and m.precipitation is not None]
        winds = [m.wind_speed for m in moscom_device_map.values() if m.device_uuid in allowed_uuids and m.wind_speed is not None]
        weather = {
            'avg_temperature': round(sum(temps)/len(temps), 1) if temps else None,
            'avg_humidity': round(sum(humids)/len(humids), 1) if humids else None,
            'avg_precipitation': round(sum(precs)/len(precs), 2) if precs else None,
            'avg_wind_speed': round(sum(winds)/len(winds), 1) if winds else None,
        }

        # 룰 기반 과학 설명 (모기 생리학 + 기상)
        explanation_rule = _build_weather_explanation(weather, signals, today_d)

        # GPT 자연어 생성 시도
        explanation_ai = None
        api_key = getattr(dj_settings, 'OPENAI_API_KEY', '') or ''
        if api_key:
            try:
                import openai
                client = openai.OpenAI(api_key=api_key)
                prompt = _forecast_gpt_prompt(weather, signals, today_d, explanation_rule)
                resp = client.chat.completions.create(
                    model='gpt-4o-mini',
                    messages=[
                        {'role': 'system', 'content': (
                            '당신은 보건소·질병청 감염병관리팀에 모기 발생 예보를 작성하는 보건학 전문가입니다. '
                            '기상 조건과 모기 생리학(우화기간, 산란, 생존율) 을 결합한 과학적 설명을 한국어로 작성하세요. '
                            '반드시 데이터에 명시된 실제 수치(기온, 습도, 강수, 모기지수, 권역명)를 인용하고, '
                            '"적정 조건", "우화 단축", "성충 증가" 등 모기 생리학 용어를 자연스럽게 사용하세요. '
                            '결과는 3~5개 문단(각 2~4문장)으로, 마지막에 행정 권고 1~3개를 불릿으로. '
                            '서론/결론 인사말 없이 바로 본문부터 시작.'
                        )},
                        {'role': 'user', 'content': prompt},
                    ],
                    temperature=0.4,
                    max_tokens=700,
                )
                explanation_ai = resp.choices[0].message.content.strip()
            except Exception as e:
                logger.warning(f'forecast_brief GPT failed: {e}')

        # ── 3. 트렌드 비교 ──
        # 최근 7일 일평균 vs 그 직전 7일 일평균
        recent7_total = 0
        recent7_count = 0
        prev7_total = 0
        prev7_count = 0
        recent7_start = today_d - timedelta(days=7)
        prev7_start = today_d - timedelta(days=14)
        for u in allowed_uuids:
            for d_str, c in daily[u].items():
                try:
                    dd = date_cls.fromisoformat(d_str)
                except ValueError:
                    continue
                if recent7_start <= dd < today_d:
                    recent7_total += c
                    recent7_count += 1
                elif prev7_start <= dd < recent7_start:
                    prev7_total += c
                    prev7_count += 1
        recent7_avg = (recent7_total / 7) if recent7_total else 0
        prev7_avg = (prev7_total / 7) if prev7_total else 0
        wow_change_pct = round((recent7_avg - prev7_avg) / prev7_avg * 100) if prev7_avg > 0 else 0

        # 작년 동기간 — 우리 Collection 에 1년치 데이터 없으면 None
        last_year_avg = None
        last_year_range = None
        if Collection is not None:
            try:
                ly_start = today_d.replace(year=today_d.year - 1) - timedelta(days=7)
                ly_end = today_d.replace(year=today_d.year - 1)
                from django.utils import timezone as dj_tz
                start_utc = dj_tz.make_aware(datetime.combine(ly_start, datetime.min.time()))
                end_utc = dj_tz.make_aware(datetime.combine(ly_end, datetime.max.time()))
                ly_qs = Collection.objects.filter(
                    device_uuid__in=allowed_uuids,
                    created_date__gte=start_utc, created_date__lte=end_utc,
                )
                ly_total = sum(c.mosquito_count for c in ly_qs)
                if ly_qs.exists():
                    last_year_avg = round(ly_total / 7, 1)
                    last_year_range = f'{ly_start.isoformat()} ~ {ly_end.isoformat()}'
            except (ValueError, Exception):
                pass

        # 작년 대비 비율
        yoy_change_pct = None
        if last_year_avg and last_year_avg > 0:
            yoy_change_pct = round((recent7_avg - last_year_avg) / last_year_avg * 100)

        trend = {
            'recent7_avg': round(recent7_avg, 1),
            'prev7_avg': round(prev7_avg, 1),
            'wow_change_pct': wow_change_pct,
            'last_year_avg': last_year_avg,
            'last_year_range': last_year_range,
            'yoy_change_pct': yoy_change_pct,
        }

        result = {
            'today': today_d.isoformat(),
            'yesterday': yday_d.isoformat(),
            'signals': signals,
            'stations': stations,
            'weather': weather,
            'explanation_rule': explanation_rule,
            'explanation_ai': explanation_ai,
            'trend': trend,
            'generated_at': datetime.now(timezone(timedelta(hours=9))).isoformat(),
        }
        cache.set(cache_key, result, 300)
        return JsonResponse(result)
    except Exception as e:
        logger.exception('forecast_brief failed')
        return JsonResponse({'error': str(e)}, status=500)


def _build_weather_explanation(weather, signals, today_d):
    """기상 → 모기 생리학 과학적 설명 (룰 기반 fallback).
    GPT 호출 실패해도 동작.
    """
    t = weather.get('avg_temperature')
    h = weather.get('avg_humidity')
    p = weather.get('avg_precipitation')
    parts = []

    # 우화 조건 평가
    if t is not None:
        if 25 <= t <= 28:
            parts.append(f'평균기온 {t}°C 는 모기 우화 최적 범위(25~28°C). 평소 우화기간 8~10일이 5~7일로 단축되어 향후 1주 이내 성충 발생량이 평소 대비 1.3~1.7배 증가할 수 있음.')
        elif 20 <= t < 25:
            parts.append(f'평균기온 {t}°C 는 우화 적정 범위 하한. 우화기간 약 8일 유지로 성충 발생은 점진적 증가 예상.')
        elif 28 < t <= 32:
            parts.append(f'평균기온 {t}°C 는 우화는 빠르나(4~5일) 28°C 초과 시 성충 활동성과 산란량은 감소. 단기적으로는 증가 후 안정세 예상.')
        elif t > 32:
            parts.append(f'평균기온 {t}°C 는 모기 활동 억제 범위. 성충 생존율 저하로 일시적 감소 가능하나, 야간 기온이 25°C 이상 유지되면 야간 활동성은 지속.')
        else:
            parts.append(f'평균기온 {t}°C 는 우화 저해 범위(20°C 미만). 우화 지연으로 성충 발생량 감소.')

    # 습도 평가
    if h is not None:
        if h >= 75:
            parts.append(f'평균 습도 {h}% 는 모기 생존·산란에 매우 유리(70% 이상). 성충 수명이 평균 14일에서 21일까지 연장될 수 있음.')
        elif h >= 60:
            parts.append(f'평균 습도 {h}% 는 모기 생존 적정. 통상적 산란 활동 유지.')
        else:
            parts.append(f'평균 습도 {h}% 는 건조 조건(60% 미만)으로 성충 수명 단축. 다만 유충 발생지에서는 영향 제한적.')

    # 강수 평가
    if p is not None and p > 0:
        if p >= 10:
            parts.append(f'최근 강수 {p}mm 는 신규 유충 발생지(고인 물) 형성에 유리. 강수 후 5~7일 사이 1차 우화 피크, 12~14일 사이 2차 피크 예상.')
        elif p >= 1:
            parts.append(f'경미한 강수({p}mm)는 기존 유충 발생지 유지 수준. 큰 변동 없을 전망.')

    # 권역별 위험 권고
    bad_regions = [s for s in signals if s['grade'] in ('주의', '불쾌')]
    if bad_regions:
        names = ', '.join(s['region_name'] for s in bad_regions[:3])
        parts.append(f'위 기상 조건 하에서 {names} 권역은 향후 3일간 모기지수 50점 이상 유지 예상. 선제 유충방제(Bti, 도시·하수구 살포) 시행 권고.')

    return '\n\n'.join(parts) if parts else '관측 데이터가 부족하여 과학적 설명을 생성할 수 없습니다.'


def _forecast_gpt_prompt(weather, signals, today_d, rule_text):
    """GPT 입력 페이로드."""
    sig_text = '\n'.join(
        f'  - {s["region_name"]} (장비 {s["device_count"]}대): 평균 모기지수 {s["avg_index"]} · {s["grade"]}'
        for s in signals[:8]
    ) or '  - 데이터 없음'
    w = weather
    return f"""[모기 발생 예보 · {today_d.isoformat()} 기준]

[1] 현재 평균 기상
  - 기온: {w.get('avg_temperature', '-')}°C
  - 습도: {w.get('avg_humidity', '-')}%
  - 강수: {w.get('avg_precipitation', 0)}mm
  - 풍속: {w.get('avg_wind_speed', '-')}m/s

[2] 향후 3일 권역별 위험도 (모기지수 0~100)
{sig_text}

[3] 룰 기반 과학적 분석 (참고용)
{rule_text}

위 데이터를 바탕으로, 모기 생리학(우화기간, 산란, 성충 생존율) 을 결합하여
"왜 모기가 증가/감소할 것인가" 를 과학적으로 설명하고,
보건소 실무자가 행정 결재 자료로 쓸 수 있도록 3~5개 문단으로 작성하세요.
마지막에는 "■ 행정 권고" 로 시작하는 불릿 권고 1~3개로 마무리.
"""


@csrf_exempt
@require_POST
def moscom_forecast_simulate(request):
    """방역 시뮬레이션 — 입력: 장비/방역방법/실시일 → 출력: 적용 전후 3일 예측 비교."""
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    try:
        import json as _json
        body = _json.loads((request.body or b'').decode('utf-8', errors='replace') or '{}')
    except Exception:
        body = {}
    device_uuid = body.get('device_uuid')
    # 다중 방역 지원: method_keys 배열 우선, 레거시 method_key fallback
    method_keys = body.get('method_keys')
    if not method_keys:
        _mk = body.get('method_key')
        method_keys = [_mk] if _mk else []
    method_keys = [m for m in (method_keys or []) if m][:2]  # 최대 2개
    apply_date = body.get('apply_date')  # YYYY-MM-DD
    if not device_uuid or not method_keys or not apply_date:
        return JsonResponse({'error': 'device_uuid, method_keys, apply_date 필요'}, status=400)
    try:
        from datetime import datetime, timedelta, timezone, date as date_cls
        apply_d = date_cls.fromisoformat(apply_date)
    except ValueError:
        return JsonResponse({'error': 'apply_date 형식 YYYY-MM-DD'}, status=400)

    su = _current_session_user(request)
    devices = moscom_client.list_devices()
    devices = user_store.filter_devices(su, devices)
    target = next((d for d in devices if d.get('device_uuid') == device_uuid), None)
    if not target:
        return JsonResponse({'error': '권한 없는 장비'}, status=403)

    # 방역 방법(들) — 각각 효과창(onset~duration) 보유
    all_methods = {m['key']: m for m in remedy_store.list_methods()}
    methods = [all_methods[k] for k in method_keys if k in all_methods]
    if not methods:
        return JsonResponse({'error': '잘못된 방역 방법'}, status=400)

    # 7일 통계
    stats = moscom_client.get_statistics(device_uuid='', period_type='2', offset=0)
    stats = [r for r in (stats or []) if r.get('device_uuid') == device_uuid]
    daily = {}
    for r in stats:
        d = (r.get('created_date') or '')[:10]
        if d:
            daily[d] = (r.get('mosquito_count') or 0)
    hist = [{'date': d, 'count': c} for d, c in sorted(daily.items())]

    # 예측 (10일치 — 방역 효과 반영 가시화 위해 길게)
    dv = target.get('device') or {}
    from core import predictor
    try:
        from moscom.models import Device as MoscomDevice
        md = MoscomDevice.objects.filter(device_uuid=device_uuid).first()
    except Exception:
        md = None

    inp = [{
        'uuid': device_uuid,
        'name': _station_name(dv.get('device_name') or '') or device_uuid,
        'region': ' '.join(p for p in [dv.get('address_sido'), dv.get('address_gungu')] if p),
        'region_code': (md.region_code if md else '') or '',
        'sido': (md.address_sido if md else '') or '',
        'history': hist,
        'weather': {
            'temperature': md.temperature if md else None,
            'humidity': md.humidity if md else None,
            'precipitation': md.precipitation if md else None,
            'wind_speed': md.wind_speed if md else None,
        } if md else {},
    }]
    preds = predictor.predict_for_devices(inp, days_ahead=10)
    if not preds:
        return JsonResponse({'error': '예측 실패'}, status=500)
    base = preds[0]

    # 방역 적용 시 예측: 방법별 효과창(apply_d + onset ~ + duration) 동안
    # 일자별로 해당 방법들의 감소율을 곱셈 누적 (adjustment_factor 와 동일 방식)
    from datetime import date as date_cls
    method_windows = []
    for m in methods:
        onset = int(m.get('onset_days', 1) or 1)
        duration = int(m.get('duration_days', 7) or 7)
        red = float(m.get('reduction_pct', 30) or 30)
        es = apply_d + timedelta(days=onset)
        ee = es + timedelta(days=duration)
        method_windows.append({
            'name': m['name'], 'key': m['key'], 'reduction_pct': red,
            'effect_start': es, 'effect_end': ee,
        })

    simulated = []
    saved_total = 0
    for p in base.get('predictions', []):
        try:
            d = date_cls.fromisoformat(p['date'])
        except ValueError:
            simulated.append(p)
            continue
        orig = p.get('predicted') or 0
        factor = 1.0
        applied_any = False
        for mw in method_windows:
            if mw['effect_start'] <= d < mw['effect_end']:
                factor *= (1 - mw['reduction_pct'] / 100.0)
                applied_any = True
        adj = max(0, int(orig * factor))
        saved_total += (orig - adj)
        simulated.append({**p, 'predicted_simulated': adj, 'effect_applied': applied_any})

    return JsonResponse({
        'device': {'uuid': device_uuid, 'name': inp[0]['name']},
        'methods': [
            {'name': mw['name'], 'key': mw['key'], 'reduction_pct': mw['reduction_pct'],
             'effect_start': mw['effect_start'].isoformat(), 'effect_end': mw['effect_end'].isoformat()}
            for mw in method_windows
        ],
        # 레거시 호환: 단일 method 필드도 첫 방법으로 유지
        'method': methods[0],
        'apply_date': apply_date,
        'predictions': simulated,
        'saved_total': saved_total,
    })


@require_GET
def moscom_hourly(request):
    """MOSCOM API raw(시간별) 데이터 프록시 — 시간별 히트맵용
    쿼리스트링: start(ISO), end(ISO) (선택, 미지정 시 전날 00:00~다음날 06:00 UTC)
    """
    auth_err = _require_mosquito_auth(request)
    if auth_err:
        return auth_err
    from datetime import datetime, timedelta, timezone
    try:
        start = request.GET.get('start')
        end = request.GET.get('end')
        if not start or not end:
            # 기본: 최근 48시간치 (수집창 18:00~05:00 커버 가능)
            now = datetime.now(timezone.utc)
            end_dt = now
            start_dt = now - timedelta(days=2)
            start = start_dt.strftime('%Y-%m-%dT%H:%M:%S.000Z')
            end = end_dt.strftime('%Y-%m-%dT%H:%M:%S.000Z')
        data = moscom_client.get_statistics_by_date(
            start_dt=start, end_dt=end, aggregation='raw', device_uuid='0',
        )
        data = _filter_records_by_uuid(request, data or [])
        return JsonResponse({
            'count': len(data) if isinstance(data, list) else 0,
            'start': start, 'end': end,
            'data': data,
        }, safe=False)
    except Exception as e:
        logger.exception('MOSCOM /device/statisticsByDate failed')
        return JsonResponse({'error': str(e)}, status=502)
